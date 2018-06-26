import json
from openpyxl import Workbook
import urllib
import boto3
import StringIO
import time
import sys, os, collections


def lambda_handler(event, context):


    s3 = boto3.client('s3')
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = urllib.unquote_plus(event['Records'][0]['s3']['object']['key']).decode('utf8')
    try:
        response = s3.get_object(Bucket=bucket, Key=key)
        data = response['Body'].read()

        var=json.loads(data)
        if 'checkpoint_export' in var:
            return export_checkpoint(var,s3,bucket)
        export_as_hotel = var.get('export_as_hotel')
        export_as_economy_view = var.get('export_as_economy_view')
        if export_as_hotel:
            return lambda_handler_room(var, s3, bucket)
        elif export_as_economy_view:
            return export_economy_view(var, s3, bucket)

        defaults=json.loads(var['defaults'])
        headers=json.loads(var['headers'])
        file_name = var['file_name']
        attendees=json.loads(var['attendees'])
        attendee_tags=json.loads(var['attendee_tags'])
        tags=json.loads(var['tags'])
        attendee_groups=json.loads(var['attendee_groups'])
        groups=json.loads(var['all_groups'])
        questions=json.loads(var['questions'])
        all_answers=json.loads(var['answers'])
        travels=json.loads(var['travels'])
        travel_list=json.loads(var['travel_list'])
        sessions=json.loads(var['sessions'])
        seminar_users=json.loads(var['seminar_users'])
        all_bookings=json.loads(var['bookings'])
        rooms=json.loads(var['rooms'])
        hotels=json.loads(var['hotels'])
        locations=json.loads(var['locations'])
        req_buddy=json.loads(var['req_buddy'])
        match_lines=json.loads(var['matchlines'])
        hotelBookings=json.loads(var['hotelBookings'])
        all_event_attendees=json.loads(var['all_event_attendees'])
        hotel_columns =var['hotel_columns']
        max_booking_number =var['max_booking_number']
        max_actual_room_buddy =var['max_actual_room_buddy']
        hotel_flag=var['hotel_flag']
        economy_columns = var['economy_columns']
        attendee_orders = json.loads(var['attendee_orders'])
        order_payments = json.loads(var['order_payments'])
        order_vat_percent_sum = json.loads(var['order_vat_percent_sum'])
        credit_usages = json.loads(var['credit_usages'])
        max_attendee_orders = var['max_attendee_orders']
        event_economy_vat_count = var['event_economy_vat_count']
        group_details = json.loads(var['group_details'])
        order_owner_excluded_list = json.loads(var['order_owner_excluded_list'])

        if hotel_flag:
            all_rows = get_hotel_sheet(all_bookings,hotelBookings,attendees,req_buddy,match_lines,rooms,hotels,questions,all_answers)
        else:
            all_rows = get_excel_body_json_format(defaults, attendees, attendee_groups, groups, tags, attendee_tags,
                                                  questions, all_answers, sessions, seminar_users, travels, travel_list,
                                                  all_bookings, hotelBookings, rooms, hotels, locations, req_buddy, match_lines,
                                                  all_event_attendees, hotel_columns, max_booking_number, max_actual_room_buddy,
                                                  economy_columns, attendee_orders, order_payments, order_vat_percent_sum, credit_usages,
                                                  max_attendee_orders, event_economy_vat_count, group_details, order_owner_excluded_list)

        all_rows = headers+all_rows

        wb = Workbook()

        excelSheet = []
        excelSheet.append(wb.active)
        excelSheet[0].title="Attendees"
        for row in all_rows:
            excelSheet[0].append(row)

        # for col in excelSheet.columns:
        #     max_length = 0
        #     column = col[0].column  # Get the column name
        #     for cell in col:
        #         try:  # Necessary to avoid error on empty cells
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(cell.value)
        #         except:
        #             pass
        #     adjusted_width = (max_length + 2) * 1.2
        #     excelSheet.column_dimensions[column].width = adjusted_width

        f = StringIO.StringIO()
        wb.save(f)
        s3.put_object(
            Bucket=event['Records'][0]['s3']['bucket']['name'],
            Key=file_name,
            Body=f.getvalue()
        )
    except Exception as e:
        # print str(e)
        # exc_type, exc_obj, exc_tb = sys.exc_info()
        # fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        # print exc_type, fname, exc_tb.tb_lineno
        print str(e)
        # print "Error"

def lambda_handler_room(var,s3,bucket):
    try:
        file_name = var['file_name']
        all_rows = export_as_room(var)
        wb = Workbook()
        ws = wb.active

        for row in all_rows:
            ws.append(row)

        f = StringIO.StringIO()
        wb.save(f)
        s3.put_object(
            Bucket=bucket,
            Key=file_name,
            Body=f.getvalue()
        )

    except Exception as e:
        # print str(e)
        # exc_type, exc_obj, exc_tb = sys.exc_info()
        # fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        # print exc_type, fname, exc_tb.tb_lineno
        print str(e)
        # print "E


def export_as_room(var):
    headers=json.loads(var['headers'])
    # attendees = json.loads(var['attendees'])
    questions = json.loads(var['questions'])
    answers = json.loads(var['answers'])
    bookings = json.loads(var['bookings'])
    matchlines = json.loads(var['matchlines'])
    rooms = json.loads(var['rooms'])

    # print(rooms)

    all_rows = headers

    for room in rooms:
        each_row=[room['hotel__name'],room['description'],room['id'],room['booking__matchline__match'],room['booking']]

        booking_id=room['booking']
        booking_details=[x for x in bookings if x['pk'] == booking_id]

        if len(booking_details)>0:
            each_row.extend([booking_details[0]['fields']['attendee'],booking_details[0]['fields']['check_in'],booking_details[0]['fields']['check_out']])
            each_row.extend(get_answer_json_format(booking_details[0]['fields']['attendee'],questions,answers))
        else:
            continue

        if room['booking__matchline__match']:
            match_list= [x for x in rooms if x['booking__matchline__match'] == room['booking__matchline__match'] and x['booking']!=room['booking']]
            if match_list:
                for match in match_list:
                    booking_id=match['booking']
                    booking_details=[x for x in bookings if x['pk'] == booking_id]
                    if booking_details:
                        each_row.extend([booking_details[0]['fields']['attendee'],booking_details[0]['fields']['check_in'],booking_details[0]['fields']['check_out']])
                        each_row.extend(get_answer_json_format(booking_details[0]['fields']['attendee'],questions,answers))

        all_rows.append(each_row)
    return all_rows


def export_economy_view(var, s3, bucket):
    try:
        file_name = var['file_name']
        all_rows = get_economy_view_data(var)

        wb = Workbook()
        ws = wb.active
        for row in all_rows:
            ws.append(row)

        f = StringIO.StringIO()
        wb.save(f)
        s3.put_object(
            Bucket=bucket,
            Key=file_name,
            Body=f.getvalue()
        )

    except Exception as e:
        print str(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print exc_type, fname, exc_tb.tb_lineno
        print str(e)

def get_economy_view_data(var):
    economy_data = []
    try:
        headers = json.loads(var['headers'])
        economy_columns = var['economy_columns']
        questions = json.loads(var['questions'])
        answers = json.loads(var['answers'])
        attendee_orders = json.loads(var['attendee_orders'])
        credit_usages = json.loads(var['credit_usages'])
        order_payments = json.loads(var['order_payments'])
        order_vat_percent_sum = json.loads(var['order_vat_percent_sum'])
        event_economy_vat_count = var['event_economy_vat_count']
        group_details = json.loads(var['group_details'])

        economy_data.append(headers)
        date_format_lambda = lambda input_date: input_date[:10] if input_date else ""
        for order in attendee_orders:
            order_data = []
            group_attendees = [int(item) for item in order['attendees'].split(',')]
            order_number = order['order_number']
            credit_usage_value = 0
            if credit_usages and credit_usages.get(order_number):
                credit_usage_value = credit_usages[order_number]
            transaction_id = ""
            transaction_date = ""
            payment_method = ""
            if order['status'] == 'paid':
                for payment in order_payments:
                    if payment['fields']['order_number'] == order_number:
                        transaction_id = payment['fields']['transaction']
                        transaction_date = date_format_lambda(payment['fields']['created_at'])
                        payment_method = 'invoice' if payment['fields']['method'] == 'admin' else 'dibs'
                        break
            if 'order-number' in economy_columns:
                order_data.append(order_number)
            if 'order-status' in economy_columns:
                order_data.append(order['status'])
            if 'invoice-id' in economy_columns:
                order_data.append(order['invoice_id'])
            if 'invoice-date' in economy_columns:
                order_data.append(date_format_lambda(order['invoice_date']))
            if 'due-date' in economy_columns:
                order_data.append(date_format_lambda(order['due_date']))
            if 'transaction-id' in economy_columns:
                order_data.append(transaction_id)
            if 'transaction-date' in economy_columns:
                order_data.append(transaction_date)
            if 'paid-by-card-invoice' in economy_columns:
                order_data.append(payment_method)
            if 'vat-xx-percent-sum' in economy_columns:
                vat_percent_sum = order_vat_percent_sum.get(order_number)
                if vat_percent_sum:
                    vat_percent_sum = collections.OrderedDict(sorted(vat_percent_sum.iteritems()))
                    for k, val in vat_percent_sum.iteritems():
                        order_data.append(val)
                else:
                    for i in range(0, event_economy_vat_count):
                        order_data.append("")
            if 'vat-total-sum' in economy_columns:
                order_data.append(order['vat_amount'])
            if 'rebate-sum' in economy_columns:
                order_data.append(order['rebate_amount'])
            if 'credit-usage' in economy_columns:
                order_data.append(credit_usage_value)
            if 'total-order-sum-excl-vat' in economy_columns:
                total_cost_excl_vat_value = order['total_cost_excl_vat'] - credit_usage_value
                order_data.append(total_cost_excl_vat_value if total_cost_excl_vat_value > 0 else 0)
            if 'total-order-sum-incl-vat' in economy_columns:
                order_data.append(order['total_cost_incl_vat'] - credit_usage_value)
            if 'order-group-id' in economy_columns:
                group_name = ""
                for  group_key, val in group_details.iteritems():
                    if group_attendees[0] in val:
                        group_name = group_key
                        break
                order_data.append(group_name)
            tt = time.time()
            for attendee in group_attendees:
                for question in questions:
                    value = ""
                    for answer in answers:
                        if attendee == answer['fields']['user'] and question['pk'] == answer['fields']['question']:
                            value = answer['fields']['value']
                            break
                    order_data.append(value)
            economy_data.append(order_data)
    except Exception as e:
        print str(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print exc_type, fname, exc_tb.tb_lineno
        print str(e)

    return economy_data


def get_excel_body_json_format(defaults, attendees, attendee_groups, groups, tags, attendee_tags, questions,
                               all_answers, sessions, seminar_users, travels, travel_list, all_bookings, hotel_bookings,
                               rooms, hotels, locations, req_buddy, match_lines, all_event_attendees, hotel_columns,
                               max_booking_number, max_actual_room_buddy, economy_columns, attendee_orders, order_payments,
                               order_vat_percent_sum, credit_usages, max_attendee_orders, event_economy_vat_count, group_details, order_owner_excluded_list):
    excel_rows=[]
    for attendee in attendees:
        each_row=[]

        if 'uid' in defaults and defaults['uid']:
            each_row.append(attendee['pk'])

        # each_row.append("No")

        if 'rdate' in defaults and defaults['rdate']:
            each_row.append(attendee['fields']['created'])
        if 'udate' in defaults and defaults['udate']:
            each_row.append(attendee['fields']['updated'])
        if 'secret' in defaults and defaults['secret']:
            each_row.append(attendee['fields']['secret_key'])
        if 'bid' in defaults and defaults['bid']:
            each_row.append(attendee['fields']['bid'])
        if 'group' in defaults and defaults['group']:
            each_row.append(get_goup_new_json_format(attendee['pk'],groups,attendee_groups))
        if 'tags' in defaults and defaults['tags']:
            each_row.append(get_tag_json_format(attendee['pk'],tags,attendee_tags))

        each_row.extend(get_answer_json_format(attendee['pk'],questions,all_answers))
        each_row.extend(add_sessions_json_format(sessions,seminar_users,attendee['pk']))
        each_row.extend(add_travels_json_format(travels,travel_list,attendee['pk']))
        each_row.extend(get_hotel_column_data(attendee['pk'], all_bookings, hotel_bookings, attendees, req_buddy, match_lines, rooms, hotels, locations, all_event_attendees, hotel_columns, max_booking_number, max_actual_room_buddy))
        each_row.extend(get_economy_column_data(attendee['pk'], economy_columns, attendee_orders, order_payments, order_vat_percent_sum, credit_usages, max_attendee_orders, event_economy_vat_count, group_details, order_owner_excluded_list))
        excel_rows.append(each_row)
    return excel_rows

def get_answer_json_format(attendee_id,questions,answers):
    # answered_questions = answers.filter(user_id=attendee_id)
    answered_questions = [x for x in answers if x['fields']['user'] == attendee_id]
    q_a_list = []
    for question in questions:
        value = ''
        for answer in answered_questions:
            if question['pk'] == answer['fields']['question']:
                value = answer['fields']['value']
        # q_a_list.append({'question_id': question['pk'], 'answer': value})
        q_a_list.append(value)
    return q_a_list


def get_tag_json_format(attendee_id,tags,all_attendee_tags):
    mytags=""
    attendee_tags= [x for x in all_attendee_tags if x['fields']['attendee'] == attendee_id]
    for tag in attendee_tags:
        my_tag_list= [x for x in tags if x['pk'] == tag['fields']['tag']]
        for my_tag in my_tag_list:
            mytags += my_tag ['fields']['name'] + ","
    if len(mytags)>0:
        mytags = mytags[:-1]
    return mytags


def get_goup_new_json_format(attendee_id,groups,all_attendee_groups):
    mygroups=""
    attendee_groups= [x for x in all_attendee_groups if x['fields']['attendee'] == attendee_id]
    for grp in attendee_groups:
        my_group_list= [x for x in groups if x['pk'] == grp['fields']['group']]
        for my_grp in my_group_list:
            mygroups += my_grp ['fields']['name'] + ","
    if len(mygroups)>0:
        mygroups = mygroups[:-1]
    return mygroups


def add_travels_json_format(all_travels,travel_users,attendee_id):
    travels=[]
    for travel in all_travels:
        traveller = [x for x in travel_users if x['fields']['attendee'] == attendee_id and x['fields']['travel']==travel['pk']]
        if traveller:
            travels.append( traveller[0]['fields']['status'])
        else:
            travels.append('')
    return travels

def add_sessions_json_format(all_sessions,session_users,attendee_id):
    sessions=[]
    for session in all_sessions:
        session_user = [x for x in session_users if x['fields']['attendee'] == attendee_id and x['fields']['session']==session['pk']]
        if session_user:
            sessions.append( session_user[0]['fields']['status'])
        else:
            sessions.append('')
    return sessions

def get_checksum(row):
    import hashlib
    # hash_object = hashlib.md5(bytes(json.dumps(row), encoding='utf-8'))
    hash_object = hashlib.md5(bytes(json.dumps(row)))
    return hash_object.hexdigest()

def get_buddy_json_format(buddy_list, attendees):
    buddies=''
    for buddy in buddy_list:
        if 'email' in buddy['fields'] and buddy['fields']['email'] :
            buddies=buddies+buddy['fields']['email']
        elif 'buddy' in buddy['fields'] and buddy['fields']['buddy']:
            attendee_info= [x for x in attendees if x['pk'] == buddy['fields']['buddy']]
            if attendee_info:
              buddies= buddies+attendee_info[0]['fields']['email']+ ","
        else :
            buddies = buddies +  ","
    return buddies

def get_hotel_sheet(bookings,hotel_all_bookings,attendees,req_buddy,matchlines,rooms,hotels,questions,all_answers):
    all_rows=[]
    for booking in bookings:
        each_row=[]
        each_row.append(booking['pk'])
        each_row.append(None)
        each_row.append(booking['fields']['attendee'])
        attendee_info= [x for x in attendees if x['pk'] == booking['fields']['attendee']]

        if attendee_info:
            each_row.append(attendee_info[0]['fields']['firstname'])
            each_row.append(attendee_info[0]['fields']['lastname'])
            each_row.append(attendee_info[0]['fields']['email'])
            room_info = [x for x in rooms if x['pk'] == booking['fields']['room']]
            hotel= [x for x in hotels if x['pk'] == room_info[0]['fields']['hotel']]
            each_row.append(booking['fields']['room'])
            each_row.append(hotel[0]['fields']['name'])
            each_row.append(room_info[0]['fields']['description'])
            each_row.append(booking['fields']['check_in'])
            each_row.append(booking['fields']['check_out'])

            requested_buddy=[x for x in req_buddy if x['fields']['booking'] == booking['pk']]
            each_row.append(get_buddy_json_format(requested_buddy,attendees))

            actualBuddyName=""
            matchIds =[x for x in matchlines if x['fields']['booking'] == booking['pk']]
            if matchIds:
                for matchId in matchIds:
                    actualBuddyList = [x for x in matchlines if x['fields']['match'] == matchId['fields']['match']]
                    for actualBuddy in actualBuddyList:
                        if actualBuddy['fields']['booking']!=booking['pk']:
                            partner_booking = [x for x in hotel_all_bookings if x['pk'] == actualBuddy['fields']['booking']]
                            if partner_booking:
                                partner_info= [x for x in attendees if x['pk'] == partner_booking[0]['fields']['attendee']]
                                if partner_info:
                                    actualBuddyName = actualBuddyName + partner_info[0]['fields']['email']+","
                                    each_row[1] = int(matchId['fields']['match'])
            if actualBuddyName:
                actualBuddyName = actualBuddyName[:-1]
            each_row.append(actualBuddyName)
            each_row.extend(get_answer_json_format(attendee_info[0]['pk'],questions,all_answers))

            all_rows.append(each_row)
    return all_rows

def get_hotel_column_data(att_id, bookings, hotel_all_bookings, attendees, req_buddy, all_matchlines, all_rooms, all_hotels, all_locations, all_event_attendees, hotel_columns, max_booking_number, max_actual_room_buddy):
    try:
        hotel_ans = []
        attendee_bookings = [x for x in bookings if x['fields']['attendee'] == att_id]
        booking_counter = 0
        for booking in attendee_bookings:
            booking_counter += 1
            room_obj = None
            hotel_obj = None
            match_ids = []
            if 'booking-id-col' in hotel_columns:
                hotel_ans.append(booking['pk'])
            if 'match-id-col' in hotel_columns:
                match_id_str = ''
                for match_line in all_matchlines:
                    if booking['pk'] == match_line['fields']['booking']:
                        match_ids.append(match_line['fields']['match'])
                        match_id_str += str(match_line['fields']['match']) + ', '
                if len(match_id_str) > 0:
                    match_id_str = match_id_str[:-2]
                hotel_ans.append(match_id_str)
            if 'room-id-col' in hotel_columns:
                # room_id = ''
                hotel_ans.append(booking['fields']['room'])
                # for room in all_rooms:
                #     if booking['fields']['room'] == room['pk']:
                #         room_obj = room
                #         room_id = room['pk']
                #         break
                # hotel_ans.append(room_id)
            if 'hotel-name-col' in hotel_columns:
                # if not room_obj:
                for room in all_rooms:
                    if booking['fields']['room'] == room['pk']:
                        room_obj = room
                        break
                hotel_name = ''
                for hotel in all_hotels:
                    if room_obj['fields']['hotel'] == hotel['pk']:
                        hotel_obj = hotel
                        hotel_name = hotel['fields']['name']
                        break
                hotel_ans.append(hotel_name)
            if 'description-col' in hotel_columns:
                room_name = ''
                if not room_obj:
                    for room in all_rooms:
                        if booking['fields']['room'] == room['pk']:
                            room_name = room['fields']['description']
                            break
                else:
                    room_name = room_obj['fields']['description']
                hotel_ans.append(room_name)
            if 'check-in-col' in hotel_columns:
                check_in = booking['fields']['check_in']
                hotel_ans.append(check_in)
            if 'check-out-col' in hotel_columns:
                check_out = booking['fields']['check_out']
                hotel_ans.append(check_out)
            if 'beds-col' in hotel_columns:
                room_beds = 0
                if not room_obj:
                    for room in all_rooms:
                        if booking['fields']['room'] == room['pk']:
                            room_beds = room['fields']['beds']
                            break
                else:
                    room_beds = room_obj['fields']['beds']
                hotel_ans.append(room_beds)
            if 'location-col' in hotel_columns:
                location_id = -1
                if hotel_obj:
                    location_id = hotel_obj['fields']['location']
                else:
                    if room_obj:
                        for hotel in all_hotels:
                            if room_obj['fields']['hotel'] == hotel['pk']:
                                location_id = hotel['fields']['location']
                                break
                    else:
                        for room in all_rooms:
                            if booking['fields']['room'] == room['pk']:
                                for hotel in all_hotels:
                                    if room['fields']['hotel'] == hotel['pk']:
                                        location_id = hotel['fields']['location']
                                        break
                                break
                location_name = ''
                for location in all_locations:
                    if location['pk'] == location_id:
                        location_name = location['fields']['name']
                        break
                hotel_ans.append(location_name)
            if 'rbr-col' in hotel_columns:
                requested_buddy = [x for x in req_buddy if x['fields']['booking'] == booking['pk']]
                hotel_ans.append(get_buddy_json_format(requested_buddy, all_event_attendees))

            if 'rba-col' in hotel_columns or 'rba-checkin-col' in hotel_columns or 'rba-checkout-col' in hotel_columns:
                # match_ids = MatchLine.objects.filter(booking_id=booking.id).values('match_id')
                if not match_ids:
                    for match_line in all_matchlines:
                        if booking['pk'] == match_line['fields']['booking']:
                            match_ids.append(match_line['fields']['match'])

                # other_bookings = MatchLine.objects.filter(match_id__in=match_ids).exclude(booking_id=booking.id)
                other_booking_ids = []
                if match_ids:
                    for other_match in all_matchlines:
                        if other_match['fields']['match'] in match_ids and booking['pk'] != other_match['fields']['booking']:
                            other_booking_ids.append(other_match['fields']['booking'])

                other_bookings = []
                if other_booking_ids:
                    other_bookings = [x for x in hotel_all_bookings if x['pk'] in other_booking_ids]

                actual_bud_counter = 0
                for other_booking in other_bookings:
                    actual_bud_counter += 1
                    other_booking_att = None
                    for att_item in all_event_attendees:
                        if att_item['pk'] == other_booking['fields']['attendee']:
                            other_booking_att = att_item
                            break
                    if 'rba-col' in hotel_columns:
                        hotel_ans.append(other_booking_att['fields']['email'])
                    if 'rba-checkin-col' in hotel_columns:
                        hotel_ans.append(other_booking['fields']['check_in'])
                    if 'rba-checkout-col' in hotel_columns:
                        hotel_ans.append(other_booking['fields']['check_out'])
                if max_actual_room_buddy > actual_bud_counter:
                    max_actual_room_buddy -= actual_bud_counter
                    for i in range(0, max_actual_room_buddy):
                        if 'rba-col' in hotel_columns:
                            hotel_ans.append('')
                        if 'rba-checkin-col' in hotel_columns:
                            hotel_ans.append('')
                        if 'rba-checkout-col' in hotel_columns:
                            hotel_ans.append('')

        if max_booking_number > booking_counter:
            max_booking_number -= booking_counter
            for i in range(0, max_booking_number):
                if 'booking-id-col' in hotel_columns:
                    hotel_ans.append('')
                if 'match-id-col' in hotel_columns:
                    hotel_ans.append('')
                if 'room-id-col' in hotel_columns:
                    hotel_ans.append('')
                if 'hotel-name-col' in hotel_columns:
                    hotel_ans.append('')
                if 'description-col' in hotel_columns:
                    hotel_ans.append('')
                if 'check-in-col' in hotel_columns:
                    hotel_ans.append('')
                if 'check-out-col' in hotel_columns:
                    hotel_ans.append('')
                if 'beds-col' in hotel_columns:
                    hotel_ans.append('')
                if 'location-col' in hotel_columns:
                    hotel_ans.append('')
                if 'rbr-col' in hotel_columns:
                    hotel_ans.append('')
                for i in range(0, max_actual_room_buddy):
                    if 'rba-col' in hotel_columns:
                        hotel_ans.append('')
                    if 'rba-checkin-col' in hotel_columns:
                        hotel_ans.append('')
                    if 'rba-checkout-col' in hotel_columns:
                        hotel_ans.append('')
        return hotel_ans
    except Exception as e:
        print 'hotel column error'
        print str(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print exc_type, fname, exc_tb.tb_lineno
        return []

def get_economy_column_data(attendee_id, economy_columns, attendee_orders, order_payments, order_vat_percent_sum, credit_usages, max_attendee_orders, event_economy_vat_count, group_details, order_owner_excluded_list):
    excel_data = []
    if not economy_columns:
        return excel_data
    try:
        order_data = dict()
        group_name = ""
        group_attendees = []
        group_details = group_details.get(str(attendee_id))
        if group_details:
            group_name = group_details['name']
            if group_details['is_owner']:
                group_attendees = group_details['members']

        order_number = ""
        # here skipped_orders and found_orders are used same purpose,
        # for group attendee,
        skipped_orders = []
        found_orders = []
        date_format_lambda = lambda input_date: input_date[:10] if input_date else ""
        for order in attendee_orders:
            if attendee_id == order['fields']['attendee'] or order_number == order['fields']['order_number'] or (group_attendees and order['fields']['attendee'] in group_attendees):
                order_number = order['fields']['order_number']
                if order_data.get(order_number):
                    order_data[order_number]['vat_total_sum'] += order['fields']['vat_amount']
                    order_data[order_number]['rebate_sum'] += order['fields']['rebate_amount']
                    order_data[order_number]['cost_excl_vat'] += order['fields']['cost']
                    order_data[order_number]['cost_incl_vat'] += order['fields']['cost'] + order['fields']['vat_amount']
                else:
                    found_orders.append(order_number)
                    transaction_id = ""
                    transaction_date = ""
                    payment_method = ""
                    if order['fields']['status'] == 'paid':
                        for payment in order_payments:
                            if payment['fields']['order_number'] == order_number:
                                transaction_id = payment['fields']['transaction']
                                transaction_date = date_format_lambda(payment['fields']['created_at'])
                                payment_method = 'invoice' if payment['fields']['method'] == 'admin' else 'dibs'
                                break
                    order_data[order_number] = dict(
                        order_number=order_number,
                        order_status=order['fields']['status'],
                        invoice_id=order['fields']['invoice_ref'],
                        invoice_date=date_format_lambda(order['fields']['invoice_date']),
                        due_date=date_format_lambda(order['fields']['due_date']),
                        transaction_id=transaction_id,
                        transaction_date=transaction_date,
                        payment_method=payment_method,
                        vat_percent_sum=order_vat_percent_sum.get(order_number),
                        vat_total_sum=order['fields']['vat_amount'],
                        rebate_sum =order['fields']['rebate_amount'],
                        cost_excl_vat=order['fields']['cost'],
                        cost_incl_vat=order['fields']['cost'] + order['fields']['vat_amount'],
                    )
            elif order_number == "":
                skipped_orders.append(order)

        for order in skipped_orders:
            if order['fields']['order_number'] in found_orders:
                if order_data.get(order['fields']['order_number']):
                    order_data[order_number]['vat_total_sum'] += order['fields']['vat_amount']
                    order_data[order_number]['rebate_sum'] += order['fields']['rebate_amount']
                    order_data[order_number]['cost_excl_vat'] += order['fields']['cost']
                    order_data[order_number]['cost_incl_vat'] += order['fields']['cost'] + order['fields']['vat_amount']

        attendee_order_counter = 0
        for key, order in order_data.iteritems():
            attendee_order_counter += 1
            excel_data = append_economy_data_to_list(attendee_id, economy_columns, excel_data, group_name, credit_usages, 0, order_owner_excluded_list, order)

        if attendee_order_counter < max_attendee_orders:
            for i in range(0, (max_attendee_orders - attendee_order_counter)):
                excel_data = append_economy_data_to_list(attendee_id, economy_columns, excel_data, group_name, credit_usages, event_economy_vat_count, order_owner_excluded_list)
    except Exception as e:
        print str(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print exc_type, fname, exc_tb.tb_lineno
    return excel_data

def append_economy_data_to_list(attendee_id, economy_columns, excel_data, group_name, credit_usages, event_economy_vat_count, order_owner_excluded_list, order=None):
    credit_usage_value = 0
    if order and credit_usages and credit_usages.get(order['order_number']):
        credit_usage_value = credit_usages[order['order_number']]
    if 'order-number' in economy_columns:
        order_number_text = ""
        if order:
            order_number_text = str(order['order_number'])
            if attendee_id not in order_owner_excluded_list:
                order_number_text += "(order owner)"
        excel_data.append(order_number_text)
    if 'order-status' in economy_columns:
        excel_data.append(order['order_status'] if order else "")
    if 'invoice-id' in economy_columns:
        excel_data.append(order['invoice_id'] if order else "")
    if 'invoice-date' in economy_columns:
        excel_data.append(order['invoice_date'] if order else "")
    if 'due-date' in economy_columns:
        excel_data.append(order['due_date'] if order else "")
    if 'transaction-id' in economy_columns:
        excel_data.append(order['transaction_id'] if order else "")
    if 'transaction-date' in economy_columns:
        excel_data.append(order['transaction_date'] if order else "")
    if 'paid-by-card-invoice' in economy_columns:
        excel_data.append(order['payment_method'] if order else "")
    if 'vat-xx-percent-sum' in economy_columns:
        if order:
            order['vat_percent_sum'] = collections.OrderedDict(sorted(order['vat_percent_sum'].iteritems()))
            for k, val in order['vat_percent_sum'].iteritems():
                excel_data.append(val if val else "")
        else:
            for i in range(0, event_economy_vat_count):
                excel_data.append("")
    if 'vat-total-sum' in economy_columns:
        excel_data.append(order['vat_total_sum'] if order else "")
    if 'rebate-sum' in economy_columns:
        excel_data.append(order['rebate_sum'] if order and order['rebate_sum'] else "")
    if 'credit-usage' in economy_columns:
        excel_data.append(credit_usage_value if credit_usage_value else "")
    if 'total-order-sum-excl-vat' in economy_columns:
        # sometime cost_excl_vat_value could be negative while credit_usage pays all amount
        cost_excl_vat_value = order['cost_excl_vat'] - credit_usage_value
        cost_excl_vat_value = cost_excl_vat_value if cost_excl_vat_value else 0
        excel_data.append(cost_excl_vat_value if order else "")
    if 'total-order-sum-incl-vat' in economy_columns:
        excel_data.append(order['cost_incl_vat'] - credit_usage_value if order else "")
    if 'order-group-id' in economy_columns:
        excel_data.append(group_name if order else "")
    return excel_data

def get_hotel_sheet_without_question(bookings,attendees,req_buddy,matchlines,rooms,hotels):
    all_rows=[]
    for booking in bookings:
        each_row=[]
        each_row.append( booking['pk'])
        each_row.append( None)
        each_row.append(booking['fields']['attendee'])
        attendee_info= [x for x in attendees if x['pk'] == booking['fields']['attendee']]

        if attendee_info:
            each_row.append(attendee_info[0]['fields']['firstname'])
            each_row.append(attendee_info[0]['fields']['lastname'])
            each_row.append(attendee_info[0]['fields']['email'])
            room_info = [x for x in rooms if x['pk'] == booking['fields']['room']]
            hotel= [x for x in hotels if x['pk'] == room_info[0]['fields']['hotel']]
            each_row.append(booking['fields']['room'])
            each_row.append(hotel[0]['fields']['name'])
            each_row.append(room_info[0]['fields']['description'])
            each_row.append(booking['fields']['check_in'])
            each_row.append(booking['fields']['check_out'])

            requested_buddy=[x for x in req_buddy if x['fields']['booking'] == booking['pk']]
            each_row.append(get_buddy_json_format(requested_buddy,attendees))



            actualBuddyName=""
            matchId =[x for x in matchlines if x['fields']['booking'] == booking['pk']]
            if matchId:
                actualBuddyList = [x for x in matchlines if x['fields']['match'] == matchId[0]['fields']['match']]
                for actualBuddy in actualBuddyList:
                    if actualBuddy['fields']['booking']!=booking['pk']:
                        partner_booking = [x for x in bookings if x['pk'] == actualBuddy['fields']['booking']]
                        if partner_booking:
                            partner_info= [x for x in attendees if x['pk'] == partner_booking[0]['fields']['attendee']]
                            if partner_info:
                                actualBuddyName = actualBuddyName + partner_info[0]['fields']['email']+","
            if actualBuddyName:
                actualBuddyName = actualBuddyName[:-1]
            each_row.append(actualBuddyName)

            all_rows.append(each_row)
    return all_rows

def export_checkpoint(var, s3, bucket):
    headers=json.loads(var['headers'])
    defaults=json.loads(var['defaults'])
    file_name = var['file_name']
    attendees = json.loads(var['attendees'])
    type = var['checkpoint_type']
    answers = json.loads(var['answers'])
    tags = json.loads(var['all_tags'])
    attendee_tags = json.loads(var['attendee_tags'])
    groups = json.loads(var['all_groups'])
    attendee_groups = json.loads(var['attendee_groups'])
    checkpoint_questions = json.loads(var['checkpoint_questions'])
    scan = json.loads(var['scan'])
    excel_rows = [headers]
    if type=='session':
        seminar_users = json.loads(var['seminar_users'])
        print('Total entry: {}'.format(len(seminar_users)))
        for seminar_user in seminar_users:
            try:
                attendee = seminar_user['fields']['attendee']
                scan_obj = [x for x in scan if x['fields']['attendee'] == attendee]
                att_info = [x for x in attendees if x['pk'] == attendee]
                if att_info:
                    att_info = att_info.pop()
                    # each_row = [att_info['fields']['secret_key'],att_info['fields']['firstname'],att_info['fields']['lastname'],att_info['fields']['email']]
                    each_row = []
                    for default in defaults:
                        if default == 'uid':
                            each_row.append(att_info['fields']['secret_key'])
                        elif default == 'bid':
                            each_row.append(att_info['fields']['bid'])
                        elif default=='created':
                            each_row.append(att_info['fields']['created'])
                        elif default=='update':
                            each_row.append(att_info['fields']['updated'])
                        elif default=='group':
                            each_row.append(get_goup_new_json_format(att_info['pk'],groups,attendee_groups))
                        elif default=='tag':
                            each_row.append(get_tag_json_format(att_info['pk'],tags,attendee_tags))

                    each_row.extend(get_answer_json_format(attendee, checkpoint_questions, answers))
                    if scan_obj:
                        scan_obj = scan_obj.pop()
                        from datetime import datetime
                        format = '%Y-%m-%dT%H:%M:%S'
                        scan_time = scan_obj['fields']['scan_time'].split('.')[0]
                        scan_time =datetime.strptime(scan_time, format)
                        each_row.extend(["IN",scan_time.time(),scan_time.date()])
                        # each_row.extend([scan_obj['fields']['scan_time'],scan_obj['fields']['scan_time'],"IN"])
                    else:
                        each_row.extend(["OUT","",""])

                excel_rows.append(each_row)

            except Exception as e:
                print str(e)
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print exc_type, fname, exc_tb.tb_lineno
    else:
        for att_info in attendees:
            try:
                scan_obj = [x for x in scan if x['fields']['attendee'] == att_info['pk']]
                # each_row = [att_info['fields']['secret_key'],att_info['fields']['firstname'],att_info['fields']['lastname'],att_info['fields']['email']]
                each_row = []

                for default in defaults:
                    if default == 'uid':
                        each_row.append(att_info['fields']['secret_key'])
                    elif default == 'bid':
                        each_row.append(att_info['fields']['bid'])
                    elif default=='created':
                        each_row.append(att_info['fields']['created'])
                    elif default=='update':
                        each_row.append(att_info['fields']['updated'])
                    elif default=='group':
                        each_row.append(get_goup_new_json_format(att_info['pk'], groups, attendee_groups))
                    elif default=='tag':
                        each_row.append(get_tag_json_format(att_info['pk'], tags, attendee_tags))

                each_row.extend(get_answer_json_format(att_info['pk'], checkpoint_questions, answers))
                if scan_obj:
                    scan_obj=scan_obj.pop()
                    from datetime import datetime
                    format = '%Y-%m-%dT%H:%M:%S'
                    scan_time = scan_obj['fields']['scan_time'].split('.')[0]
                    scan_time =datetime.strptime(scan_time, format)
                    each_row.extend(["IN",scan_time.time(),scan_time.date()])
                    # each_row.extend([scan_obj['fields']['scan_time'],scan_obj['fields']['scan_time'],"IN"])
                else:
                    each_row.extend(["OUT","",""])

                excel_rows.append(each_row)

            except Exception as e:
                print str(e)
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print exc_type, fname, exc_tb.tb_lineno

    try:
        wb = Workbook()
        ws = wb.active

        for row in excel_rows:
            ws.append(row)

        f = StringIO.StringIO()
        wb.save(f)
        s3.put_object(
            Bucket=bucket,
            Key=file_name,
            Body=f.getvalue()
        )
    except Exception as e:
        print str(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print exc_type, fname, exc_tb.tb_lineno
