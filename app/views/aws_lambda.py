import json
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import urllib
import boto3
# import StringIO
import time


import io
import os
from django.http import HttpResponse



def export(var):

    var=json.loads(var)

    if 'checkpoint_export' in var:
        return export_checkpoint(var)

    export_as_hotel = var['export_as_hotel']
    if export_as_hotel:
        return export_as_room(var)



    defaults=json.loads(var['defaults'])
    headers=json.loads(var['headers'])
    file_name = var['file_name']
    attendees=json.loads(var['attendees'])
    attendee_tags=json.loads(var['attendee_tags'])
    tags=json.loads(var['tags'])
    groups=json.loads(var['all_groups'])
    attendee_groups=json.loads(var['attendee_groups'])
    questions=json.loads(var['questions'])
    all_answers=json.loads(var['answers'])
    travels=json.loads(var['travels'])
    travel_list=json.loads(var['travel_list'])
    sessions=json.loads(var['sessions'])
    seminar_users=json.loads(var['seminar_users'])
    all_bookings=json.loads(var['bookings'])
    rooms=json.loads(var['rooms'])
    hotels=json.loads(var['hotels'])
    locations=json.loads(var['locations'])
    req_buddy=json.loads(var['req_buddy'])
    match_lines=json.loads(var['matchlines'])
    hotelBookings=json.loads(var['hotelBookings'])
    hotel_flag=var['hotel_flag']

    if hotel_flag:
        all_rows = get_hotel_sheet(all_bookings,hotelBookings,attendees,req_buddy,match_lines,rooms,hotels,questions,all_answers)
    else:
        all_rows= get_excel_body_json_format(defaults,attendees,attendee_groups,groups,tags,attendee_tags,questions,all_answers,sessions,seminar_users,travels,travel_list,all_bookings,rooms,hotels,locations,req_buddy,match_lines)


    all_rows = headers+all_rows

    wb = Workbook()

    excelSheet = []
    excelSheet.append(wb.active)
    excelSheet[0].title="Attendees"
    for row in all_rows:
        excelSheet[0].append(row)

    # if hotel_flag:
    #     hotel_sheet=[]
    #     hotel_headers=['Attendee Id', 'First Name', 'Last Name', 'Email', 'Room Description', 'Hotel Name', 'Check In', 'Check Out', 'Room Buddy Requested', 'Room Buddy Actual']
    #     hotel_sheet.append(hotel_headers)
    #     hotel_sheet.extend(get_hotel_sheet(all_bookings,attendees,req_buddy,match_lines,rooms,hotels,questions,all_answers))
    #
    #     ws2=wb.create_sheet()
    #     ws2.title="Hotels"
    #     ws2.append(hotel_headers)
    #     for hotel in hotel_sheet:
    #         ws2.append(hotel)
    #     excelSheet.append(ws2)



    # if not os.path.exists("attendeeList"):
    #     os.makedirs("attendeeList")
    f = io.BytesIO()
    wb.save(f)


    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="AllAttendee.xls"'
    return response


def lambda_handler(event, context):


    s3 = boto3.client('s3')
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = urllib.unquote_plus(event['Records'][0]['s3']['object']['key']).decode('utf8')
    try:
        response = s3.get_object(Bucket=bucket, Key=key)
        data = response['Body'].read()

        var=json.loads(data)

        if 'checkpoint_export' in var:
            return export_checkpoint(var)

        export_as_hotel = var['export_as_hotel']
        if export_as_hotel:
            return lambda_handler_room(var,s3,bucket)


        defaults=json.loads(var['defaults'])
        headers=json.loads(var['headers'])
        file_name = var['file_name']
        attendees=json.loads(var['attendees'])
        attendee_tags=json.loads(var['attendee_tags'])
        tags=json.loads(var['tags'])
        groups=json.loads(var['all_groups'])
        attendee_groups=json.loads(var['attendee_groups'])
        questions=json.loads(var['questions'])
        all_answers=json.loads(var['answers'])
        travels=json.loads(var['travels'])
        travel_list=json.loads(var['travel_list'])
        sessions=json.loads(var['sessions'])
        seminar_users=json.loads(var['seminar_users'])
        all_bookings=json.loads(var['bookings'])
        rooms=json.loads(var['rooms'])
        hotels=json.loads(var['hotels'])
        locations=json.loads(var['locations'])
        req_buddy=json.loads(var['req_buddy'])
        match_lines=json.loads(var['matchlines'])
        hotelBookings=json.loads(var['hotelBookings'])
        hotel_flag=var['hotel_flag']

        if hotel_flag:
            all_rows = get_hotel_sheet(all_bookings,attendees,req_buddy,match_lines,rooms,hotels,questions,all_answers)
        else:
            all_rows= get_excel_body_json_format(defaults,attendees,attendee_groups,groups,tags,attendee_tags,questions,all_answers,sessions,seminar_users,travels,travel_list,all_bookings,rooms,hotels,locations,req_buddy,match_lines)


        all_rows = headers+all_rows

        wb = Workbook()

        excelSheet = []
        excelSheet.append(wb.active)
        excelSheet[0].title="Attendees"
        for row in all_rows:
            excelSheet[0].append(row)

        # f = StringIO.StringIO()
        # wb.save(f)
        # s3.put_object(
        #     Bucket=event['Records'][0]['s3']['bucket']['name'],
        #     Key=file_name,
        #     Body=f.getvalue()
        # )
    except Exception as e:
        # print str(e)
        print(str(e))

def lambda_handler_room(var,s3,bucket):
    try:

        file_name = var['file_name']

        all_rows=export_as_room(var)
        wb = Workbook()
        ws = wb.active

        for row in all_rows:
            ws.append(row)

        # f = StringIO.StringIO()
        # wb.save(f)
        # s3.put_object(
        #     Bucket=bucket,
        #     Key=file_name,
        #     Body=f.getvalue()
        # )


    except Exception as e:
        # print str(e)
        print(str(e))


def export_as_room(var):
    headers=json.loads(var['headers'])
    file_name = var['file_name']
    # attendees = json.loads(var['attendees'])
    questions = json.loads(var['questions'])
    answers = json.loads(var['answers'])
    bookings = json.loads(var['bookings'])
    matchlines = json.loads(var['matchlines'])
    rooms = json.loads(var['rooms'])

    # print(rooms)

    all_rows = headers

    for room in rooms:
        each_row=[room['hotel__name'],room['description'],room['id'],room['booking__matchline__match'],room['booking']]

        booking_id=room['booking']
        booking_details=[x for x in bookings if x['pk'] == booking_id]

        if len(booking_details)>0:
            each_row.extend([booking_details[0]['fields']['attendee'],booking_details[0]['fields']['check_in'],booking_details[0]['fields']['check_out']])
            each_row.extend(get_answer_json_format(booking_details[0]['fields']['attendee'],questions,answers))
        else:
            continue

        if room['booking__matchline__match']:
            match_list= [x for x in rooms if x['booking__matchline__match'] == room['booking__matchline__match'] and x['booking']!=room['booking']]
            if match_list:
                for match in match_list:
                    booking_id=match['booking']
                    booking_details=[x for x in bookings if x['pk'] == booking_id]
                    if booking_details:
                        each_row.extend([booking_details[0]['fields']['attendee'],booking_details[0]['fields']['check_in'],booking_details[0]['fields']['check_out']])
                        each_row.extend(get_answer_json_format(booking_details[0]['fields']['attendee'],questions,answers))

        all_rows.append(each_row)
        # else:
        #     each_row.extend([booking_details[0]['fields']['attendee'],booking_details[0]['fields']['check_in'],booking_details[0]['fields']['check_out']])
        #     each_row.extend(get_answer_json_format(booking_details[0]['fields']['attendee'],questions,answers))
        #     all_rows.append(each_row)

    print("come2")
    return all_rows
    # wb = Workbook()
    # ws = wb.active
    # #
    # for row in all_rows:
    #     ws.append(row)
    # #
    # if not os.path.exists("attendeeList"):
    #     os.makedirs("attendeeList")
    # f = io.BytesIO()
    # wb.save(f)
    #
    # response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    # response['Content-Disposition'] = 'attachment; filename="AllAttendee.xls"'
    # return response




def export_as_room_old(var):
    headers=json.loads(var['headers'])
    file_name = var['file_name']
    # attendees = json.loads(var['attendees'])
    questions = json.loads(var['questions'])
    answers = json.loads(var['answers'])
    bookings = json.loads(var['bookings'])
    matchlines = json.loads(var['matchlines'])
    rooms = json.loads(var['rooms'])

    print("come 1")

    all_rows = headers

    for room in rooms:
        room = json.loads(room)
        match_list= [x for x in matchlines if x['fields']['match'] == room['match_id']]
        # match_list= [x for x in matchlines if x['pk'] == 1 ]

        if match_list:
            each_row=[room['booking__room__hotel__name'],room['booking__room__description']]
            for matchline in match_list:
                booking_id=matchline['fields']['booking']
                booking_details=[x for x in bookings if x['pk'] == booking_id]
                # print(booking_details)
                each_row.extend([booking_details[0]['fields']['attendee'],booking_details[0]['fields']['check_in'],booking_details[0]['fields']['check_out']])
                each_row.extend(get_answer_json_format(booking_details[0]['fields']['attendee'],questions,answers))
            all_rows.append(each_row)
    print("come2")
    return all_rows
    # wb = Workbook()
    # ws = wb.active
    # #
    # for row in all_rows:
    #     ws.append(row)
    # #
    # if not os.path.exists("attendeeList"):
    #     os.makedirs("attendeeList")
    # f = io.BytesIO()
    # wb.save(f)
    #
    # response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    # response['Content-Disposition'] = 'attachment; filename="AllAttendee.xls"'
    # return response




def get_excel_body_json_format(defaults,attendees,attendee_groups,groups,tags,attendee_tags,questions,all_answers,sessions,seminar_users,travels,travel_list,all_bookings,rooms,hotels,locations,req_buddy,match_lines):
    excel_rows=[]
    for attendee in attendees:
        each_row=[]

        if 'uid' in defaults and defaults['uid']:
            each_row.append(attendee['pk'])

        each_row.append("No")

        if 'rdate' in defaults and defaults['rdate']:
            each_row.append(attendee['fields']['created'])
        if 'udate' in defaults and defaults['udate']:
            each_row.append(attendee['fields']['updated'])
        if 'secret' in defaults and defaults['secret']:
            each_row.append(attendee['fields']['secret_key'])
        if 'group' in defaults and defaults['group']:
            each_row.append(get_goup_new_json_format(attendee['pk'],groups,attendee_groups))
        if 'tags' in defaults and defaults['tags']:
            each_row.append(get_tag_json_format(attendee['pk'],tags,attendee_tags))


        each_row.extend(get_answer_json_format(attendee['pk'],questions,all_answers))
        each_row.extend(add_sessions_json_format(sessions,seminar_users,attendee['pk']))
        each_row.extend(add_travels_json_format(travels,travel_list,attendee['pk']))

        # row = each_row[:]
        # row.pop(1)
        # row.pop(2)
        # # print(row)
        # checksum=get_checksum(row)
        # if attendee['fields']['checksum'] == checksum:
        #     each_row[1] = ""
        # else:
        #     each_row[1] = "No"

        excel_rows.append(each_row)
    return excel_rows


def get_answer_json_format(attendee_id,questions,answers):
    # answered_questions = answers.filter(user_id=attendee_id)
    answered_questions = [x for x in answers if x['fields']['user'] == attendee_id]
    q_a_list = []
    for question in questions:
        value = ''
        for answer in answered_questions:
            if question['pk'] == answer['fields']['question']:
                value = answer['fields']['value']
        # q_a_list.append({'question_id': question['pk'], 'answer': value})
        q_a_list.append(value)
    return q_a_list


def get_tag_json_format(attendee_id,tags,all_attendee_tags):
    mytags=""
    attendee_tags= [x for x in all_attendee_tags if x['fields']['attendee'] == attendee_id]
    for tag in attendee_tags:
        my_tag_list= [x for x in tags if x['pk'] == tag['fields']['tag']]
        for my_tag in my_tag_list:
            mytags += my_tag ['fields']['name'] + ","
    if len(mytags)>0:
        mytags = mytags[:-1]
    return mytags

def get_goup_new_json_format(attendee_id,groups,all_attendee_groups):
    mygroups=""
    attendee_groups= [x for x in all_attendee_groups if x['fields']['attendee'] == attendee_id]
    for grp in attendee_groups:
        my_group_list= [x for x in groups if x['pk'] == grp['fields']['group']]
        for my_grp in my_group_list:
            mygroups += my_grp ['fields']['name'] + ","
    if len(mygroups)>0:
        mygroups = mygroups[:-1]
    return mygroups

def get_group_json_format(group_id,groups):
    group_name=""
    group = [x for x in groups if x['pk'] == group_id]
    if group:
        group_name=group[0]['fields']['name']
    return group_name


def add_travels_json_format(all_travels,travel_users,attendee_id):
    travels=[]
    for travel in all_travels:
        traveller = [x for x in travel_users if x['fields']['attendee'] == attendee_id and x['fields']['travel']==travel['pk']]
        if traveller:
            travels.append( traveller[0]['fields']['status'])
        else:
            travels.append('')
    return travels

def add_sessions_json_format(all_sessions,session_users,attendee_id):
    sessions=[]
    for session in all_sessions:
        session_user = [x for x in session_users if x['fields']['attendee'] == attendee_id and x['fields']['session']==session['pk']]
        if session_user:
            sessions.append( session_user[0]['fields']['status'])
        else:
            sessions.append('')
    return sessions

def get_checksum(row):
    import hashlib
    hash_object = hashlib.md5(bytes(json.dumps(row), encoding='utf-8'))
    return hash_object.hexdigest()

def get_buddy_json_format(buddy_list,attendees):
    buddies=''
    for buddy in buddy_list:
        if 'email' in buddy['fields'] and buddy['fields']['email'] :
            buddies=buddies+buddy['fields']['email']
        elif 'buddy' in buddy['fields'] and buddy['fields']['buddy']:
            attendee_info= [x for x in attendees if x['pk'] == buddy['fields']['buddy']]
            if attendee_info:
              buddies= buddies+attendee_info[0]['fields']['email']+ " "
        else :
            buddies = buddies+ " "
    return buddies

def get_hotel_sheet(bookings,hotel_all_bookings,attendees,req_buddy,matchlines,rooms,hotels,questions,all_answers):
    all_rows=[]
    for booking in bookings:
        each_row=[]
        each_row.append(booking['pk'])
        each_row.append(None)
        each_row.append(booking['fields']['attendee'])
        attendee_info= [x for x in attendees if x['pk'] == booking['fields']['attendee']]

        if attendee_info:
            each_row.append(attendee_info[0]['fields']['firstname'])
            each_row.append(attendee_info[0]['fields']['lastname'])
            each_row.append(attendee_info[0]['fields']['email'])
            room_info = [x for x in rooms if x['pk'] == booking['fields']['room']]
            hotel= [x for x in hotels if x['pk'] == room_info[0]['fields']['hotel']]
            each_row.append(booking['fields']['room'])
            each_row.append(hotel[0]['fields']['name'])
            each_row.append(room_info[0]['fields']['description'])
            each_row.append(booking['fields']['check_in'])
            each_row.append(booking['fields']['check_out'])

            requested_buddy=[x for x in req_buddy if x['fields']['booking'] == booking['pk']]
            each_row.append(get_buddy_json_format(requested_buddy,attendees))

            actualBuddyName=""
            matchIds =[x for x in matchlines if x['fields']['booking'] == booking['pk']]
            # print (matchlines)
            if matchIds:
                for matchId in matchIds:
                    actualBuddyList = [x for x in matchlines if x['fields']['match'] == matchId['fields']['match']]
                    for actualBuddy in actualBuddyList:
                        if actualBuddy['fields']['booking']!=booking['pk']:
                            partner_booking = [x for x in hotel_all_bookings if x['pk'] == actualBuddy['fields']['booking']]
                            if partner_booking:
                                partner_info= [x for x in attendees if x['pk'] == partner_booking[0]['fields']['attendee']]
                                if partner_info:
                                    actualBuddyName = actualBuddyName + partner_info[0]['fields']['email']+","
                                    each_row[1] = int(matchId['fields']['match'])
            if actualBuddyName:
                actualBuddyName = actualBuddyName[:-1]
            each_row.append(actualBuddyName)
            each_row.extend(get_answer_json_format(attendee_info[0]['pk'],questions,all_answers))

            all_rows.append(each_row)
    return all_rows

def get_hotel_sheet_without_question(bookings,attendees,req_buddy,matchlines,rooms,hotels):
    all_rows=[]
    for booking in bookings:
        each_row=[]
        each_row.append(booking['pk'])
        each_row.append(None)
        each_row.append(booking['fields']['attendee'])
        attendee_info= [x for x in attendees if x['pk'] == booking['fields']['attendee']]

        if attendee_info:
            each_row.append(attendee_info[0]['fields']['firstname'])
            each_row.append(attendee_info[0]['fields']['lastname'])
            each_row.append(attendee_info[0]['fields']['email'])
            room_info = [x for x in rooms if x['pk'] == booking['fields']['room']]
            hotel= [x for x in hotels if x['pk'] == room_info[0]['fields']['hotel']]
            each_row.append(booking['fields']['room'])
            each_row.append(hotel[0]['fields']['name'])
            each_row.append(room_info[0]['fields']['description'])
            each_row.append(booking['fields']['check_in'])
            each_row.append(booking['fields']['check_out'])

            requested_buddy=[x for x in req_buddy if x['fields']['booking'] == booking['pk']]
            each_row.append(get_buddy_json_format(requested_buddy,attendees))



            actualBuddyName=""
            matchId =[x for x in matchlines if x['fields']['booking'] == booking['pk']]
            if matchId:
                actualBuddyList = [x for x in matchlines if x['fields']['match'] == matchId[0]['fields']['match']]
                for actualBuddy in actualBuddyList:
                    if actualBuddy['fields']['booking']!=booking['pk']:
                        partner_booking = [x for x in bookings if x['pk'] == actualBuddy['fields']['booking']]
                        if partner_booking:
                            partner_info= [x for x in attendees if x['pk'] == partner_booking[0]['fields']['attendee']]
                            if partner_info:
                                actualBuddyName = actualBuddyName + partner_info[0]['fields']['email']+","
            if actualBuddyName:
                actualBuddyName = actualBuddyName[:-1]
            each_row.append(actualBuddyName)

            all_rows.append(each_row)
    return all_rows


def export_checkpoint(var):
    # var=json.loads(var)
    # print(var)

    headers=json.loads(var['headers'])
    defaults=json.loads(var['defaults'])
    file_name = var['file_name']
    attendees = json.loads(var['attendees'])
    type = var['checkpoint_type']
    answers = json.loads(var['answers'])
    tags = json.loads(var['all_tags'])
    attendee_tags = json.loads(var['attendee_tags'])
    groups = json.loads(var['all_groups'])
    attendee_groups = json.loads(var['attendee_groups'])
    checkpoint_questions = json.loads(var['checkpoint_questions'])
    scan = json.loads(var['scan'])
    excel_rows = [headers]
    if type=='session':
        seminar_users = json.loads(var['seminar_users'])
        print(len(seminar_users))
        for seminar_user in seminar_users:
            attendee = seminar_user['fields']['attendee']
            scan_obj = [x for x in scan if x['fields']['attendee'] == attendee]
            att_info = [x for x in attendees if x['pk'] == attendee]
            if att_info:
                att_info = att_info.pop()
                each_row = [att_info['fields']['secret_key'],att_info['fields']['firstname'],att_info['fields']['lastname'],att_info['fields']['email']]
                each_row.extend(get_answer_json_format(attendee,checkpoint_questions,answers))

                for default in defaults:
                    if default=='created':
                        each_row.append(att_info['fields']['created'])
                    if default=='update':
                        each_row.append(att_info['fields']['updated'])
                    if default=='group':
                        each_row.append(get_goup_new_json_format(att_info['pk'],groups,attendee_groups))
                    if default=='tag':
                        each_row.append(get_tag_json_format(att_info['pk'],tags,attendee_tags))

                if scan_obj:
                    scan_obj = scan_obj.pop()
                    from datetime import datetime
                    format = '%Y-%m-%dT%H:%M:%S'
                    scan_time =datetime.strptime(scan_obj['fields']['scan_time'],format )
                    each_row.extend(["IN",scan_time.time(),scan_time.date()])
                    # each_row.extend([scan_obj['fields']['scan_time'],scan_obj['fields']['scan_time'],"IN"])
                else:
                    each_row.extend(["OUT","",""])
            excel_rows.append(each_row)
            print(each_row)
            # print(scan_obj.pop())
            # return HttpResponse("sdfsdf")
    else:
        for att_info in attendees:
            scan_obj = [x for x in scan if x['fields']['attendee'] == att_info['pk']]
            each_row = [att_info['fields']['secret_key'],att_info['fields']['firstname'],att_info['fields']['lastname'],att_info['fields']['email']]
            each_row.extend(get_answer_json_format(att_info['pk'],checkpoint_questions,answers))

            for default in defaults:
                if default=='created':
                    each_row.append(att_info['fields']['created'])
                if default=='update':
                    each_row.append(att_info['fields']['updated'])
                if default=='group':
                    each_row.append(get_goup_new_json_format(att_info['pk'],groups,attendee_groups))
                if default=='tag':
                    each_row.append(get_tag_json_format(att_info['pk'],tags,attendee_tags))

            if scan_obj:
                scan_obj=scan_obj.pop()
                from datetime import datetime
                format = '%Y-%m-%dT%H:%M:%S'
                scan_time =datetime.strptime(scan_obj['fields']['scan_time'],format )
                each_row.extend(["IN",scan_time.time(),scan_time.date()])
                # each_row.extend([scan_obj['fields']['scan_time'],scan_obj['fields']['scan_time'],"IN"])
            else:
                each_row.extend(["OUT","",""])
            # print(each_row)
            excel_rows.append(each_row)
            # print(scan_obj.pop())
    wb = Workbook()
    excelSheet = []
    excelSheet.append(wb.active)
    excelSheet[0].title="Checkpoints"
    for row in excel_rows:
        excelSheet[0].append(row)
    f = io.BytesIO()
    wb.save(f)

    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename='+file_name
    return response

    # try:
    #     wb = Workbook()
    #     ws = wb.active
    #
    #     for row in excel_rows:
    #         ws.append(row)
    #
    #     f = StringIO.StringIO()
    #     wb.save(f)
    #     s3.put_object(
    #         Bucket=bucket,
    #         Key=file_name,
    #         Body=f.getvalue()
    #     )
    # except Exception as e:
    #     print str(e)


