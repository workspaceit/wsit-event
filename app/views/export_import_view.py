
try:
    unicode = unicode
except NameError:
    # 'unicode' is undefined, must be Python 3
    str = str
    unicode = str
    bytes = bytes
    basestring = (str, bytes)
else:
    # 'unicode' exists, must be Python 2
    str = str
    unicode = unicode
    bytes = str
    basestring = basestring

from django.shortcuts import render, redirect

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import os, threading, time
from django.template.loader import render_to_string

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views import generic
from django.contrib.auth.hashers import make_password
from app.models import Attendee, MatchLine, RoomAllotment, Match, Group, Scan, ExportRule, SessionRating, RuleSet, \
    Session, AttendeeTag, \
    SeminarSpeakers, SessionTags, Events, SeminarsUsers, Questions, ExportState, Booking, Answers, RequestedBuddy, \
    TravelAttendee, Travel, \
    ImportChangeRequest, Option, ActivityHistory, Setting, Tag, AttendeeGroups, ImportChangeStatus, EmailContents, \
    SessionClasses, MessageHistory

import json
import logging
import string
import random
from django.db.models import Q
from django.db import transaction, connection
from .filter import FilterView
from .common_views import GroupView, EventView, CommonContext, TimeDetailView
from .hotel_view import HotelView
from .session_views import SessionView
from django.views.decorators.http import require_http_methods
from django.conf import settings
from boto3.session import Session as boto_session
from datetime import datetime, timedelta
from django.db.models import Max
import re
import io
from .export_lambda import ExcelView as ev
from .general_view import General
from .lambda_import import LambdaUpload
from publicfront.views.profile import SessionDetail
from pytz import timezone
from.email_view import EmailView
from .language_view import LanguageView
from app.views.gbhelper.error_report_helper import ErrorR
from app.views.lambda_import import HotelLamda
from app.views.gbhelper.language_helper import LanguageH
from app.views.gbhelper.economy_library import EconomyLibrary


class ExcelView(generic.DetailView):
    allAnswers = []
    seminarUsers = []
    sessionFilterings = []

    def exported_files_list(request):
        session = boto_session(aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                               aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                               region_name='eu-west-1')
        client = session.client('s3')
        time_with_timezone = ExcelView.getTimezoneNow(request)
        time_difference_positive = str(time_with_timezone).split("+")
        time_difference_negetive = str(time_with_timezone).split("-")
        if len(time_difference_positive) > 1:
            hours = time_difference_positive[1].split(':')[0]
        elif len(time_difference_negetive) > 3:
            hours = '-' + time_difference_negetive[3].split(':')[0]
        else:
            hours = 0
        hours = int(hours)
        event = Events.objects.get(id=request.session['event_auth_user']['event_id'])
        response = client.list_objects(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Prefix='exported_files/' + event.name + '/'
        )
        newlist = []
        if 'Contents' in response:
            from operator import itemgetter
            newlist = sorted(response['Contents'], key=itemgetter('LastModified'), reverse=True)
            for export_file in newlist:
                export_file['file_name'] = export_file['Key'].split('/')[2]
                export_file['LastModified'] = export_file['LastModified'] + timedelta(hours=hours)

        context = {
            'file_list': newlist
        }
        return render(request, 'export-import/file_list.html', context)

    def export_state(request):
        response_data = {}
        response_data['msg'] = []
        response_data['next_ajax_req'] = False
        now = datetime.now()
        time_before_5min = now - timedelta(minutes=7)
        export_state = ExportState.objects.filter(admin_id=request.session['event_auth_user']['id'],
                                                  event_id=request.session['event_auth_user']['event_id'], status=0,
                                                  created__gt=time_before_5min)
        if export_state.exists():
            session = boto_session(aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                                   aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                                   region_name='eu-west-1')
            client = session.client('s3')
            for each_file in export_state:
                key = each_file.file_name
                try:
                    response = client.get_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=key)
                    # print(response)
                    if response:
                        from django.core.urlresolvers import reverse
                        filename_arr = key.split('/')
                        if len(filename_arr) > 2:
                            filename_key = filename_arr[2]
                        else:
                            filename_key = filename_arr[1]
                        key = key.replace('&', 'and_char_checker')
                        msg = [{'filename': key, 'message': "Your file is ready!<a class='notification-link' href='" + reverse('downloadExportedFile') + "?export=" + key + "'>Click here to download</a>"}]
                        response_data['msg'].extend(msg)
                        each_file.status = 1
                        each_file.save()
                    else:
                        response_data['next_ajax_req'] = True
                except:
                    response_data['next_ajax_req'] = True
        else:
            response_data['next_ajax_req'] = False
        return HttpResponse(json.dumps(response_data), content_type="application/json")

    def download_from_s3(request):
        try:
            session = boto_session(aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                                   aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                                   region_name='eu-west-1')
            client = session.client('s3')
            key = request.GET.get('export')
            filename_arr = key.split('/')
            if len(filename_arr) > 2:
                filename = filename_arr[2]
            else:
                filename = filename_arr[1]
            filename = filename.replace('and_char_checker', '&')
            key = key.replace('and_char_checker', '&')
            response = client.get_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=key)
            f = response['Body'].read()
            response = HttpResponse(f, content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="' + filename + '"'
            return response
        except Exception as exc:
            ErrorR.efail(exc)
            return HttpResponse('Error occurred to download file.')


    def delete_from_s3(request):
        session = boto_session(aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                               aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                               region_name='eu-west-1')
        client = session.client('s3')
        response = client.delete_object(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=request.POST.get('key')
        )
        response_data = {'success': 'success'}
        return HttpResponse(json.dumps(response_data), content_type="application/json")

    def get_answer_new(attendee_id, questions, sessionFlag):
        q_list = []
        global allAnswers
        answered_questions = allAnswers.filter(user_id=attendee_id)

        for question in questions:
            value = ''
            for answer in answered_questions:
                if question.id == answer.question_id:
                    value = answer.value
            q_list.append({'question_id': question.id, 'answer': value})

        if sessionFlag:
            sessions = []
            global seminarUsers
            sessilonList = seminarUsers.filter(attendee_id=attendee_id, status='attending')
            for x in range(0, 6):
                if len(sessions) < 6:
                    sessions.append({'question_id': "session", 'answer': ""})
                else:
                    sessions = sessions[:6]
            q_list.extend(sessions)

        return q_list

    def export_hotel(request):
        if 'is_login' not in request.session or not request.session['is_login']:
            return redirect('login')
        else:
            header = ['UID', 'Title', 'First Name', 'Last Name', 'Email', 'Room Buddy Request', 'Room Buddy Actual',
                      'Check In', 'Check Out', 'Office', 'Passport Number', 'Date of Expire', 'Nationality',
                      'Country of issue of passport', 'Country of residence', 'Date of Birth']

            wb = Workbook()
            ws = wb.active
            ws.append(header)
            allBookings = Booking.objects.all()
            for booking in allBookings:
                body_row = []
                body_row.append(booking.attendee_id)

                answered_questions = Answers.objects.filter(user_id=booking.attendee_id)

                title = answered_questions.filter(question_id=36)
                if title:
                    body_row.append(title[0].value)
                else:
                    body_row.append('')

                room_buddy = answered_questions.filter(question_id=107)

                body_row.append(booking.attendee.firstname)
                body_row.append(booking.attendee.lastname)
                body_row.append(booking.attendee.email)

                buddyList = RequestedBuddy.objects.filter(booking_id=booking.id)
                buddies = ''
                for buddy in buddyList:
                    if buddy.name:
                        buddies += buddy.name
                    elif buddy.buddy_id:
                        buddies += buddy.buddy.firstname + " " + buddy.buddy.lastname
                    else:
                        buddies = ""

                body_row.append(buddies)
                actualBuddyName = ''
                if room_buddy.exists():
                    actualBuddyName = room_buddy[0].value
                else:
                    matchId = MatchLine.objects.filter(booking_id=booking.id)
                    if matchId:
                        actualBuddyList = MatchLine.objects.filter(match_id=matchId[0].match_id)
                        for actualBuddy in actualBuddyList:
                            if actualBuddy.booking_id != booking.id:
                                actualBuddyName = actualBuddy.booking.attendee.firstname + " " + actualBuddy.booking.attendee.lastname

                body_row.append(actualBuddyName)

                body_row.append(booking.check_in)
                body_row.append(booking.check_out)

                office = answered_questions.filter(question_id=56)
                if office:
                    body_row.append(office[0].value)
                else:
                    body_row.append('')

                passport = answered_questions.filter(question_id=42)
                if passport:
                    body_row.append(passport[0].value)
                else:
                    body_row.append('')

                dateExpire = answered_questions.filter(question_id=47)
                if dateExpire:
                    body_row.append(dateExpire[0].value)
                else:
                    body_row.append('')

                nationality = answered_questions.filter(question_id=44)
                if nationality:
                    body_row.append(nationality[0].value)
                else:
                    body_row.append('')

                passportCountry = answered_questions.filter(question_id=43)
                if passportCountry:
                    body_row.append(passportCountry[0].value)
                else:
                    body_row.append('')

                residenceCountry = answered_questions.filter(question_id=49)
                if residenceCountry:
                    body_row.append(residenceCountry[0].value)
                else:
                    body_row.append('')

                dob = answered_questions.filter(question_id=41)
                if dob:
                    body_row.append(dob[0].value)
                else:
                    body_row.append('')

                ws.append(body_row)

            if not os.path.exists("HotelList"):
                os.makedirs("HotelList")

            wb.save("HotelList/Hotels.xlsx")
            response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="Hotels.xls"'
            return response

    def allExportList(request):
        if 'is_login' not in request.session or not request.session['is_login']:
            return redirect('login')
        else:
            global allAnswers
            allAnswers = Answers.objects.all()
            global seminarUsers
            seminarUsers = SeminarsUsers.objects.all()
            allData = []
            qArr = ['UID', 'Registration Date', 'Update Date']
            question_groups = GroupView.get_questionGroup(request)
            questions = []
            for group in question_groups:
                questions.extend(Questions.objects.filter(group=group))
            for question in questions:
                if question.title.lower() == "room buddy":
                    continue

                qArr.append(question.title)

            sessionHeaders = ['Outbound flight', 'Outbound flight departue date & time',
                              'Outbound flight arrival date & time', 'Homebound fight',
                              'Homebound flight departue date & time', 'Homebound flight arrival date & time']
            qArr.extend(sessionHeaders)
            bookingHeaders = ['Room Buddy Request', 'Room Buddy Actual']
            qArr.extend(bookingHeaders)

            attendees = Attendee.objects.filter(status="registered")
            wb = Workbook()
            ws = wb.active
            ws.append(qArr)
            reqBuddy = RequestedBuddy.objects.all()
            matchLines = MatchLine.objects.all()
            allBookings = Booking.objects.all()
            for attendee in attendees:
                question_list = ExcelView.get_answer_new(attendee.id, questions, True)

                attRow = [attendee.id, attendee.created, attendee.updated]
                i = 1
                actualBuddyName = ''
                for qqq in question_list:
                    i = i + 1
                    if qqq['question_id'] == 107:
                        actualBuddyName = qqq['answer']
                        continue
                    attRow.append(qqq['answer'])

                attendeeBookings = allBookings.filter(attendee_id=attendee.id)
                buddies = ''

                if actualBuddyName == '':
                    for booking in attendeeBookings:

                        buddyList = reqBuddy.filter(booking_id=booking.id)

                        for buddy in buddyList:
                            if buddy.name:
                                buddies = buddies + buddy.name
                            elif buddy.buddy_id:
                                buddies = buddies + buddy.buddy.firstname + " " + buddy.buddy.lastname
                            else:
                                buddies = buddies + ""

                        matchId = matchLines.filter(booking_id=booking.id)

                        if matchId:
                            actualBuddyList = matchLines.filter(match_id=matchId[0].match_id)
                            for actualBuddy in actualBuddyList:
                                if actualBuddy.booking_id != booking.id:
                                    actualBuddyName = actualBuddyName + actualBuddy.booking.attendee.firstname + " " + actualBuddy.booking.attendee.lastname

                attRow.append(buddies)
                attRow.append(actualBuddyName)
                ws.append(attRow)

            if not os.path.exists("attendeeList"):
                os.makedirs("attendeeList")

            wb.save("attendeeList/AllAttendee.xlsx")
            response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="AllAttendee.xls"'

            return response

    def saveAnswers(userId, answer):
        x_answer = Answers.objects.filter(question_id=answer['question_id'], user_id=userId)

        if answer['answer'] == '[delete]':
            if answer['question_id'] == 63 or answer['question_id'] == 64 or answer['question_id'] == 65:
                return "Can not delete. Required Filed "
            else:
                x_answer.delete()
                return "success"

        attendeeAnswerValidation = ExcelView.checkValidation(answer['a_type'], answer['answer'])
        if attendeeAnswerValidation != "success":
            return attendeeAnswerValidation

        if x_answer:
            attendeeAnswer = x_answer.update(value=answer['answer'])
        else:
            attendeeAnswer = Answers(value=answer['answer'], question_id=answer['question_id'], user_id=userId)
            attendeeAnswer.save()
        return attendeeAnswerValidation

    def saveTravel(userId, travel):
        try:
            x_travel = TravelAttendee.objects.filter(travel_id=travel['travel_id'], attendee_id=userId)
            if travel['status'] == '[delete]':
                x_travel.delete()
                return "success"

            if x_travel:
                attendee_travel = x_travel.update(status=travel['status'])
            else:
                attendee_travel = TravelAttendee(status=travel['status'], travel_id=travel['travel_id'],
                                                 attendee_id=userId)
                attendee_travel.save()
            return "success"
        except Exception as e:
            return str(e)

    def checkValidation(fieldType, value):
        if fieldType == 'text':
            message = 'success'
        elif fieldType == 'phone':
            message = 'success'
        elif fieldType == 'email':
            if ExcelView.validateEmail(value) == 1:
                message = 'success'
            else:
                message = "Email is not in the correct fromat"

        elif fieldType == 'date':
            message = 'success'
            # if EmailTryView.validate_date(value)==1:
            #     message='success'
            # else:
            #     message="Date is not in the correct fromat"
        elif fieldType == 'time':
            message = 'success'
        else:
            message = "Unknown Type"
        return message

    def validateEmail(email):

        if len(email) > 7:
            if re.match("^.+\\@(\\[?)[a-zA-Z0-9\\-\\.]+\\.([a-zA-Z]{2,3}|[0-9]{1,3})(\\]?)$", email) != None:
                return 1
        return 0

    def validate_mobile(value):
        if len(value) > 5:
            # if re.match("'^(?:\+?44)?[07]\d{9,13}$'", value) != None:
            #     return 1

            rule = re.compile(r'(^[+0-9]{1,3})*([0-9]{10,11}$)')
            if rule.match(value):
                return 1
        return 0

    def validate_date(value):
        try:
            f = '%Y-%m-%d'
            datetime.strptime(value, f)
            return 1
        except:
            return 0

    def export_session(request):

        if 'sessions' in request.POST:
            session_ids = request.POST.get('sessions')
            session_ids = session_ids.split(",")
            sessions = Session.objects.filter(id__in=session_ids)
        else:
            sessions = Session.objects.all()
        return ExcelView.export_session_details(request, sessions)

    def export_session_by_id(request, pk):
        sessions = Session.objects.filter(id=pk)
        ssn_name = "Session Attending User.xls"
        if sessions:
            ssn_name = sessions[0].name + ".xlsx"
        return ExcelView.export_session_details(request, sessions, ssn_name)

    def export_session_details(request, allSession, filename="Session Attending User.xls"):

        header = ['Session Id', 'Session Name', 'Attendee Id']

        visible_questions = SessionView.get_visible_questions(request)
        visible_columns_info = SessionView.get_visible_column_info(visible_questions)
        header.extend(visible_columns_info)

        wb = Workbook()

        excelSheet = [[]]
        excelSheet.append(wb.active)
        user_id = request.session['event_auth_user']['id']
        event_id = request.session['event_auth_user']['event_id']
        firstname = ''
        lastname = ''
        email = ''
        phone = ''
        office = ''

        try:

            i = 0
            for session in allSession:

                sName = session.name
                sName = re.sub('[^A-Za-z0-9 ]+', '', sName)
                if len(sName) > 30:
                    sName = sName[:30]
                if i > 0:
                    excelSheet.append(wb.create_sheet(i, sName))
                elif i == 0:
                    excelSheet[1].title = sName
                i = i + 1
                excelSheet[i].append(header)
                sessionId = session.id
                session = SeminarsUsers.objects.filter(session_id=sessionId, session__group__event_id=event_id,
                                                       status='attending')
                SessionView.get_all_attendee_visible_info(session, visible_questions)
                questions = Questions.objects.filter(group__event_id=event_id, actual_definition__isnull=False)

                for ssn in session:
                    att_answer = Answers.objects.filter(user_id=ssn.attendee_id, question_id__in=questions)
                    visible_answers = []
                    for answer in ssn.question_answers:
                        visible_answers.append(answer['answer'])
                    excelSheet[i].append(
                        [sessionId, ssn.session.name, ssn.attendee.id] + visible_answers)

            f = io.BytesIO()
            wb.save(f)
            response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=' + filename
            return response

        except Exception as e:
            return HttpResponse(json.dumps(['Exception ' + str(e)]), content_type="application/json")

    def convertBooleanToYesNo(value):
        if value:
            return "Yes"
        else:
            return "No"

    def get_session_speakers(sid):
        speakers = ""
        session_speakers = SeminarSpeakers.objects.filter(session_id=sid)
        for speaker in session_speakers:
            speakers += speaker.speaker.email + ", "
        if len(speakers) > 0:
            speakers = speakers[:-2]
        return speakers

    def get_session_tags(sid):
        tags = ""
        session_tag = SessionTags.objects.filter(session_id=sid)
        for tag in session_tag:
            tags += tag.tag.name + ", "
        if len(tags) > 0:
            tags = tags[:-2]
        return tags

    def get_session_classes(sid):
        classes = ""
        session_classes = SessionClasses.objects.filter(session_id=sid)
        for class_name in session_classes:
            classes += class_name.classname.classname + ", "
        if len(classes) > 0:
            classes = classes[:-2]
        return classes

    def get_session_vat(session):
        if session.vat:
            return session.vat
        else:
            return ''

    def export_secret(request):

        header = ['UID', 'Email', 'First Name', 'Last Name', 'Office', 'First Name Badge', 'Last Name Badge']
        allAnswers = Answers.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.append(header)

        try:

            allAtt = Attendee.objects.filter(status="registered")

            for att in allAtt:
                fName = ''
                lName = ''
                email = ''
                ofc = ''
                fNameBadge = ''
                lNameBadge = ''
                secret = att.secret_key
                myAnswers = Answers.objects.filter(Q(user_id=att.id) & (
                    Q(question_id=63) | Q(question_id=64) | Q(question_id=65) | Q(question_id=56) | Q(
                        question_id=68) | Q(
                        question_id=69)))
                for ans in myAnswers:
                    if ans.question_id == 63:
                        fName = ans.value
                    if ans.question_id == 64:
                        lName = ans.value
                    if ans.question_id == 65:
                        email = ans.value
                    if ans.question_id == 56:
                        ofc = ans.value
                    if ans.question_id == 68:
                        fNameBadge = ans.value
                    if ans.question_id == 69:
                        lNameBadge = ans.value

                ws.append([secret, email, fName, lName, ofc, fNameBadge, lNameBadge])

            if not os.path.exists("attendeeList"):
                os.makedirs("attendeeList")

            wb.save("attendeeList/User Secrets.xlsx")
            response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="User Secrets.xls"'
            return response

        except Exception as e:
            return HttpResponse(json.dumps(['Exception ' + str(e)]), content_type="application/json")

    def export_scan_list(request):
        try:
            logger = logging.getLogger(__name__)
            test_start = datetime.now()
            var = "teSt two"
            for x in range(0, 1000000):
                if var.lower() == "test two":
                    var2 = ""

            logger.debug('Testing takes time ' + str(datetime.now() - test_start))

            start_time = datetime.now()
            logger.debug('start function. Time ' + str(start_time))

            header = ['User Id', 'User Secret', 'First Name', 'Last Name', 'Time', 'office', 'email']
            event_id = request.session['event_auth_user']['event_id']
            scan_list = Scan.objects.filter(attendee__event_id=event_id)
            rows = [header]

            logger.debug('Total Scan Object ' + str(len(scan_list)))

            indx = 0
            for scan in scan_list:
                loop_start_time = datetime.now()
                ofc = ''
                email = ''
                fName = ''
                lName = ''

                fName_ans = Answers.objects.filter(user_id=scan.attendee_id, question__actual_definition='firstname',
                                                   question__group__event_id=event_id)
                if fName_ans:
                    fName = fName_ans[0].value
                lName_ans = Answers.objects.filter(user_id=scan.attendee_id, question__actual_definition='lastname',
                                                   question__group__event_id=event_id)
                if lName_ans:
                    lName = lName_ans[0].value
                email_ans = Answers.objects.filter(user_id=scan.attendee_id, question__actual_definition='email',
                                                   question__group__event_id=event_id)
                if email_ans:
                    email = email_ans[0].value
                office_ans = Answers.objects.filter(user_id=scan.attendee_id, question__actual_definition='office',
                                                    question__group__event_id=event_id)
                if office_ans:
                    ofc = office_ans[0].value

                rows.append([scan.attendee_id, scan.attendee.secret_key, fName, lName, scan.scan_time, ofc, email])

                if indx % 1000 == 0:
                    logger.debug('indx ' + str(indx) + " taking " + str(datetime.now() - loop_start_time))
                    # logger.debug('Total Answer Object '+ str(len(myAnswers)))
                indx += 1

            row_making_time = datetime.now()
            logger.debug('Row prepared finished. Time taken ' + str(row_making_time - start_time))
            response = ev.write_excel(rows, "Scan List.xlsx")
            response_time = datetime.now()

            logger.debug('write excel time taken ' + str(response_time - row_making_time))

            # S3
            wb = Workbook()
            ws = wb.active
            file_name = "scanlist/Scan List.xlsx"

            for row in rows:
                ws.append(row)

            f = io.BytesIO()
            wb.save(f)

            import boto
            from boto.s3.key import Key
            conn = boto.connect_s3(settings.AWS_ACCESS_KEY_ID, settings.AWS_SECRET_ACCESS_KEY, host=settings.AWS_STORAGE_HOST)
            bucket = conn.get_bucket(settings.AWS_STORAGE_BUCKET_NAME)
            key_name = file_name
            k = Key(bucket)
            k.key = key_name

            if not k.exists():
                key = bucket.new_key(key_name)
                key.set_contents_from_string(f.getvalue())
                key.set_acl('public-read')
                key.make_public()
            else:
                k.set_contents_from_string(f.getvalue())
                k.set_acl('public-read')
                k.make_public()

            export_state = ExportState(file_name=file_name, status=0,
                                       event_id=request.session['event_auth_user']['event_id'],
                                       admin_id=request.session['event_auth_user']['id'])
            export_state.save()

            logger.debug('Total time taken ' + str(response_time - start_time))
            return JsonResponse({'msg':"Scan List export is processing. You will be notified after finishing process."})

        except Exception as e:
            ErrorR.efail(e)
            return HttpResponse(json.dumps(['Exception ' + str(e)]), content_type="application/json")

    def export_ratings(request):

        header = ['Attendee Id', 'First Name', 'Last Name', 'Email', 'Session Name', 'Session Id', 'Ratings']

        wb = Workbook()
        ws = wb.active
        ws.append(header)

        try:
            rating_list = SessionRating.objects.filter(
                session__group__event_id=request.session['event_auth_user']['event_id'])

            for rating in rating_list:
                ofc = ''
                email = ''
                fName = ''
                lName = ''

                myAnswers = Answers.objects.filter(
                    Q(user_id=rating.attendee_id) & (Q(question_id=63) | Q(question_id=64) | Q(question_id=65)))
                for ans in myAnswers:
                    if ans.question_id == 63:
                        fName = ans.value
                    if ans.question_id == 64:
                        lName = ans.value
                    if ans.question_id == 65:
                        email = ans.value

                    if fName == '':
                        fName = rating.attendee.firstname
                    if lName == '':
                        lName = rating.attendee.lastname
                    if email == '':
                        email = rating.attendee.email

                ws.append(
                    [rating.attendee_id, fName, lName, email, rating.session.name, rating.session.id, rating.rating])

            if not os.path.exists("ratingList"):
                os.makedirs("ratingList")

            wb.save("ratingList/Rating List.xlsx")
            response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="Rating List.xls"'
            return response

        except Exception as e:
            return HttpResponse(json.dumps(['Exception ' + str(e)]), content_type="application/json")

    def get_modal_html(request):
        updateFlag = False
        export_types = [{'value': 'attendee_edit', 'checked': 'checked', 'label': 'Export file for Attendee editing'},
                        {'value': 'hotel_view', 'checked': '', 'label': 'Export file for Hotel viewing'},
                        {'value': 'economy_view', 'checked': '', 'label': 'Export file for Economy viewing'}]

        hotel_columns = ''
        economy_columns = ''
        include_import_header = False
        if 'id' in request.POST:
            updateFlag = True
            current_id = request.POST.get('id')
            current_rule = ExportRule.objects.get(id=current_id)
            data = json.loads(current_rule.preset)
            current_qlist = data['questions'].split(',')
            current_slist = data['sessions'].split(',')
            current_rule_id = data['rule_id']
            current_uid = data['uid']
            current_rdate = data['rdate']
            current_udate = data['udate']
            current_secret = data['secret']
            current_bid = data.get('bid')
            current_attGroup = data['attGroup']
            current_attTag = data['attTag']
            current_hotel = data['hotel']
            current_flight = data['flight']
            include_import_header = 'checked' if data.get('include_import_header') else ""
            if 'hotel_columns' in data:
                hotel_columns = data.get('hotel_columns')
            if 'economy_columns' in data:
                economy_columns = data.get('economy_columns')

            if 'export_type' in data:
                current_type = data['export_type']
                for export_type in export_types:
                    if export_type['value'] == current_type:
                        export_types[0]['checked'] = ''
                        export_type['checked'] = 'checked'
        groups = GroupView.get_exportfilterGroup(request)
        if updateFlag:
            for group in groups:
                if group.id == current_rule.group.id:
                    group.selected = 'selected'
        questions_group = GroupView.get_questionGroup(request)

        questions = []
        for group in questions_group:
            questions.extend(Questions.objects.filter(group=group).order_by('question_order'))

        filterGroup = ExcelView.get_filterGroup(request)
        for fff in filterGroup:
            for ff in fff.filters:
                if updateFlag == True and str(ff.id) == str(current_rule_id):
                    ff.selected = 'selected'

        session_group = GroupView.get_sessionGroup(request)
        sessionList = []
        for group in session_group:
            sessionList.extend(Session.objects.filter(group=group).order_by('session_order'))
        crnt = -1
        for ssn in sessionList:
            if ssn.group_id != crnt:
                ssn.newGroup = True
                crnt = ssn.group_id
            else:
                ssn.newGroup = False
            if (updateFlag == True) and str(ssn.id) in current_slist:
                ssn.checked = 'checked'

        hotel_column_list = [
            dict(name='Booking Id', id='booking-id-col', checked='checked' if 'booking-id-col' in hotel_columns else ''),
            dict(name='Match Id', id='match-id-col', checked='checked' if 'match-id-col' in hotel_columns else ''),
            dict(name='Hotel name', id='hotel-name-col', checked='checked' if 'hotel-name-col' in hotel_columns else ''),
            dict(name='Description', id='description-col', checked='checked' if 'description-col' in hotel_columns else ''),
            dict(name='Room Id', id='room-id-col', checked='checked' if 'room-id-col' in hotel_columns else ''),
            dict(name='Check-in', id='check-in-col', checked='checked' if 'check-in-col' in hotel_columns else ''),
            dict(name='Check-out', id='check-out-col', checked='checked' if 'check-out-col' in hotel_columns else ''),
            dict(name='Beds', id='beds-col', checked='checked' if 'beds-col' in hotel_columns else ''),
            dict(name='Location', id='location-col', checked='checked' if 'location-col' in hotel_columns else ''),
            dict(name='Room buddy requested', id='rbr-col', checked='checked' if 'rbr-col' in hotel_columns else ''),
            dict(name='Room buddy actual', id='rba-col', checked='checked' if 'rba-col' in hotel_columns else ''),
            dict(name='Room buddy check-in', id='rba-checkin-col', checked='checked' if 'rba-checkin-col' in hotel_columns else ''),
            dict(name='Room buddy check-out', id='rba-checkout-col', checked='checked' if 'rba-checkout-col' in hotel_columns else ''),
        ]
        economy_column_list = [
            dict(name='Order Number', id='order-number', checked='checked' if 'order-number' in economy_columns else ''),
            dict(name='Order Status', id='order-status', checked='checked' if 'order-status' in economy_columns else ''),
            dict(name='Invoice ID', id='invoice-id', checked='checked' if 'invoice-id' in economy_columns else ''),
            dict(name='Invoice Created Date', id='invoice-date', checked='checked' if 'invoice-date' in economy_columns else ''),
            dict(name='Due Date', id='due-date', checked='checked' if 'due-date' in economy_columns else ''),
            dict(name='Transaction ID', id='transaction-id', checked='checked' if 'transaction-id' in economy_columns else ''),
            dict(name='Transaction Date', id='transaction-date', checked='checked' if 'transaction-date' in economy_columns else ''),
            dict(name='Paid by Card / Invoice', id='paid-by-card-invoice', checked='checked' if 'paid-by-card-invoice' in economy_columns else ''),
            dict(name='VAT XX% Sum', id='vat-xx-percent-sum', checked='checked' if 'vat-xx-percent-sum' in economy_columns else ''),
            dict(name='VAT Total Sum', id='vat-total-sum', checked='checked' if 'vat-total-sum' in economy_columns else ''),
            dict(name='Rebate Sum', id='rebate-sum', checked='checked' if 'rebate-sum' in economy_columns else ''),
            dict(name='Credit Usage', id='credit-usage', checked='checked' if 'credit-usage' in economy_columns else ''),
            dict(name='Total Order Sum excl. VAT', id='total-order-sum-excl-vat', checked='checked' if 'total-order-sum-excl-vat' in economy_columns else ''),
            dict(name='Total Order Sum incl. VAT', id='total-order-sum-incl-vat', checked='checked' if 'total-order-sum-incl-vat' in economy_columns else ''),
            dict(name='Order Group ID', id='order-group-id', checked='checked' if 'order-group-id' in economy_columns else ''),
        ]


        crnt = -1
        for qqq in questions:
            if qqq.group_id != crnt:
                qqq.newGroup = True
                crnt = qqq.group_id
            else:
                qqq.newGroup = False

            if (updateFlag == True) and str(qqq.id) in current_qlist:
                qqq.checked = 'checked'

        allq = []

        # DEFAULT GROUP
        registrattionDate = {'newGroup': True, 'id': 'registrattionDate', 'title': 'Registration Date',
                             'group': {'name': 'General'}}
        if updateFlag and current_rdate:
            registrattionDate.update({'checked': 'checked'})
        allq.append(registrattionDate)
        updateDate = {'newGroup': False, 'id': 'updateDate', 'title': 'Update Date', 'group': {'name': 'General'}}
        if updateFlag and current_udate:
            updateDate.update({'checked': 'checked'})
        allq.append(updateDate)
        uid = {'newGroup': False, 'id': 'uid', 'title': 'UID', 'group': {'name': 'General'}}
        if updateFlag and current_uid:
            uid.update({'checked': 'checked'})
        allq.append(uid)
        secret = {'newGroup': False, 'id': 'secret', 'title': 'UID External', 'group': {'name': 'General'}}
        if updateFlag and current_secret:
            secret.update({'checked': 'checked'})
        allq.append(secret)
        bid = {'newGroup': False, 'id': 'bid', 'title': 'BID Badge', 'group': {'name': 'General'}}
        if updateFlag and current_bid:
            bid.update({'checked': 'checked'})
        allq.append(bid)
        attGroup = {'newGroup': False, 'id': 'attGroup', 'title': 'Attendee Group', 'group': {'name': 'General'}}
        if updateFlag and current_attGroup:
            attGroup.update({'checked': 'checked'})
        allq.append(attGroup)
        attTag = {'newGroup': False, 'id': 'attTag', 'title': 'Attendee Tag', 'group': {'name': 'General'}}
        if updateFlag and current_attTag:
            attTag.update({'checked': 'checked'})
        allq.append(attTag)
        # END DEFAULT GROUP

        allq.extend(questions)

        context = {
            'questions': allq,
            'filterGroup': filterGroup,
            'sessionList': sessionList,
            'groups': groups,
            'export_types': export_types,
            'hotel_column_list': hotel_column_list,
            'economy_column_list': economy_column_list,
            'include_import_header': include_import_header
        }
        if updateFlag:
            context.update({'name': current_rule.name})
            if current_hotel:
                context.update({'hotels': 'checked'})
            if current_flight:
                context.update({'flight': 'checked'})

        modal_html = render_to_string('export-import/export_filter_edit.html', context)
        response_data = {}
        response_data['modal_html'] = modal_html
        return HttpResponse(json.dumps(response_data), content_type="application/json")

    def export_filters(request):
        if 'is_login' not in request.session or not request.session['is_login']:
            return redirect('login')
        else:
            if request.method == 'GET':
                if EventView.check_read_permissions(request, 'export_filter_permission'):
                    exportfilterGroup = GroupView.get_exportfilterGroup(request)
                    for group in exportfilterGroup:
                        group.filters = ExportRule.objects.filter(group_id=group.id).order_by('export_order')

                    common_context = CommonContext.get_all_common_context(request)
                    context = {
                        'exportFilterGroup': exportfilterGroup,
                    }
                    context.update(common_context)
                    filter_context= CommonContext.get_filter_context(request)
                    context.update(filter_context)
                    return render(request, 'export-import/export.html', context)
            else:
                if EventView.check_permissions(request, 'export_filter_permission'):
                    q_list = request.POST.getlist('questions')
                    slist = request.POST.getlist('sessions')
                    rule_id = request.POST.get('rule_id')
                    group_id = request.POST.get('group_id')
                    filter_name = request.POST.get('filter_name')
                    export_type = request.POST.get('export_type')
                    hotel_columns = request.POST.getlist('hotel_columns')
                    hotel_columns = ','.join(map(str, hotel_columns)) if hotel_columns else ''
                    economy_columns = request.POST.getlist('economy_columns')
                    economy_columns = ','.join(map(str, economy_columns)) if economy_columns else ''
                    include_import_header = True if request.POST.get('include_import_header') else False

                    uid = True if request.POST.get('uid') else False
                    rDate = True if request.POST.get('registrattionDate') else False
                    uDate = True if request.POST.get('updateDate') else False
                    secret = True if request.POST.get('secret') else False
                    bid = True if request.POST.get('bid') else False
                    attGroup = True if request.POST.get('attGroup') else False
                    attTag = True if request.POST.get('attTag') else False
                    hotel = True if request.POST.get('hotels') else False
                    flight = True if request.POST.get('flight') else False

                    preset = {'questions': ','.join(map(str, q_list)), 'sessions': ','.join(map(str, slist)),
                              'rule_id': rule_id, 'uid': uid, 'rdate': rDate, 'udate': uDate, 'secret': secret, 'bid': bid,
                              'attGroup': attGroup, 'attTag': attTag, 'hotel': hotel, 'flight': flight,
                              'export_type': export_type, 'hotel_columns': hotel_columns, 'economy_columns': economy_columns,
                              'include_import_header': include_import_header}

                    if 'id' in request.POST:
                        id = request.POST.get('id')
                        old_rule = ExportRule.objects.filter(pk=id)
                        old_group = old_rule[0].group.id
                        old_rule.update(preset=json.dumps(preset), group_id=group_id, name=filter_name)
                        return HttpResponse(json.dumps(
                            {'id': id, 'name': filter_name, 'group_id': group_id, 'update': 'success',
                             'old_group': old_group}))
                    else:
                        export_order = ExcelView.get_export_order(group_id)

                        er = ExportRule(name=filter_name, preset=json.dumps(preset), group_id=group_id,
                                        created_by_id=request.session['event_auth_user']['id'],
                                        export_order=export_order)
                        er.save()
                        return HttpResponse(json.dumps({'id': er.id, 'name': filter_name, 'group_id': group_id}))
                else:
                    return HttpResponse(json.dumps({'error': 'You do not have Permission to do this'}))

    def search(request):
        search_key = request.POST.get('search_key')
        all_filters_groups = []
        if search_key:
            filters_group = Group.objects.filter(
                Q(type="export_filter", is_show=1, event_id=request.session['event_auth_user']['event_id']) & (
                    Q(exportrule__name__icontains=search_key))).order_by(
                'group_order').distinct()
        else:
            filters_group = Group.objects.filter(
                Q(type="export_filter", is_show=1, event_id=request.session['event_auth_user']['event_id'])).order_by(
                'group_order').distinct()

        for group in filters_group:
            group.filters = ExportRule.objects.filter(Q(group_id=group.id) & Q(name__icontains=search_key)).order_by(
                'export_order')
            group_dict = dict(
                id=group.id,
                name=group.name,
                filters=group.filters
            )
            all_filters_groups.append(group_dict)
        data = {
            'filterGroup': all_filters_groups
        }
        return render(request, 'export-import/export_filter_result.html', data)

    def get_filterGroup(request):
        filterGroup = GroupView.get_filterGroup(request)
        for group in filterGroup:
            group.filters = RuleSet.objects.all().filter(group_id=group.id).exclude(name='quick-filter')
        return filterGroup

    def filter_duplicate(request):
        response_data = {}
        event_id = request.session['event_auth_user']['event_id']
        if EventView.check_permissions(request, 'export_filter_permission'):
            filter_id = request.POST.get('filter_id')
            filter = ExportRule.objects.get(id=filter_id)

            duplicate_existance = ExportRule.objects.filter(name=filter.name + '[Copy]', group__event_id=event_id)
            if duplicate_existance.exists():
                response_data['error'] = 'This export is already make a duplicate.'
                return HttpResponse(json.dumps(response_data), content_type='application/json')

            filter.pk = None
            # if '[Copy]' not in session.name:
            filter.name += '[Copy]'
            filter.created_by_id = request.session['event_auth_user']['id']
            filter.save()
            response_data['success'] = "Create duplicate filter Successfully"
            response_data['filter'] = filter.as_dict()
        else:
            response_data['error'] = 'You do not have Permission to do this'
        return HttpResponse(json.dumps(response_data), content_type='application/json')

    def delete(request):
        response_data = {}
        if EventView.check_permissions(request, 'export_filter_permission'):
            id = request.POST.get('id')
            rule = ExportRule.objects.get(id=id)
            rule.delete()
            response_data['success'] = 'Filter Preset Deleted Successfully'
        else:
            response_data['error'] = 'You do not have Permission to do this'
        return HttpResponse(json.dumps(response_data), content_type="application/json")

    @require_http_methods(["POST"])
    def export(request):
        id = request.POST.get('id')
        rule = ExportRule.objects.get(id=id)
        data = json.loads(rule.preset)
        export_filter_name = rule.name
        qlist = data['questions'].split(',')
        slist = data['sessions'].split(',')
        hotel_columns = None
        economy_columns = None
        include_import_header = data.get('include_import_header')
        if 'hotel_columns' in data:
            hotel_columns = data.get('hotel_columns')
        if 'economy_columns' in data:
            economy_columns = data.get('economy_columns')
        if len(data['questions']) < 1:
            qlist = []
        if len(data['sessions']) < 1:
            slist = []

        export_type = "attendee_edit"
        if "export_type" in data:
            if data['export_type'] in ['hotel_view', 'economy_view']:
                export_type = data['export_type']
        return ev.export(request, export_type, export_filter_name, qlist, slist, data['rule_id'], data['uid'], data['rdate'],
                         data['udate'], data['secret'], data.get('bid'), data['attGroup'], data['attTag'], data['hotel'], data['flight'],
                         hotel_columns, economy_columns, include_import_header)

    def all_export_filter(request, qlist, slist, rule_id, uid, rdate, udate, secret, attGroup, attTag, hotel, travel):
        attendees = FilterView.get_filtered_attendees(request, rule_id)
        global allAnswers
        allAnswers = Answers.objects.filter(question_id__in=qlist, user__in=attendees)
        global seminarUsers
        seminarUsers = SeminarsUsers.objects.filter(attendee__in=attendees)
        qArr = []
        if uid:
            qArr.append("UID")
        if rdate:
            qArr.append("Registration Date")
        if udate:
            qArr.append("Update Date")
        if secret:
            qArr.append("UID (External)")
        if attGroup:
            qArr.append("Attendee Group")
        if attTag:
            qArr.append("Attendee Tag")
            allAttendeeTags = AttendeeTag.objects.filter(attendee__in=attendees)
        question_groups = GroupView.get_questionGroup(request)
        questions = []
        for group in question_groups:
            questions.extend(Questions.objects.filter(group=group, id__in=qlist))
        for question in questions:
            if question.title.lower() == "room buddy":
                continue
            qArr.append(question.title)
        if travel:
            travelHeaders = ['Outbound flight', 'Outbound flight departue date & time',
                             'Outbound flight arrival date & time', 'Homebound fight',
                             'Homebound flight departue date & time', 'Homebound flight arrival date & time']
            qArr.extend(travelHeaders)
        global sessionFilterings
        sessionFilterings = Session.objects.filter(id__in=slist)
        for sssn in sessionFilterings:
            qArr.append(sssn.name)

        if hotel:
            bookingHeaders = ['Hotel Name', 'Hotel Location', 'Room Buddy Request', 'Room Buddy Actual', 'Beds', 'Cost']
            qArr.extend(bookingHeaders)
            reqBuddy = RequestedBuddy.objects.all()
            matchLines = MatchLine.objects.all()
            allBookings = Booking.objects.all()

        wb = Workbook()
        ws = wb.active
        ws.append(qArr)

        for attendee in attendees:
            question_list = ExcelView.get_answer_new(attendee.id, questions, travel)
            attRow = []
            if uid:
                attRow.append(attendee.id)

            if rdate:
                attRow.append(attendee.created)

            if udate:
                attRow.append(attendee.updated)

            if secret:
                attRow.append(attendee.secret_key)

            if attGroup:
                attRow.append('Attending')

            if attTag:
                allAttendeeTags.filter(attendee_id=attendee.id)
                taglist = ''
                for tag in allAttendeeTags:
                    taglist = taglist + tag.tag.name + ","
                    taglist = taglist[:-1]
                attRow.append(taglist)

            i = 1
            actualBuddyName = ''
            for qqq in question_list:
                i = i + 1
                if qqq['question_id'] == 107:
                    actualBuddyName = qqq['answer']
                    continue
                attRow.append(qqq['answer'])

            for sssn in sessionFilterings:
                session_status = seminarUsers.filter(session_id=sssn.id, attendee_id=attendee.id)
                if session_status:
                    attRow.append(session_status[0].status)
                else:
                    attRow.append("Not-Answered")

            if hotel:
                attendeeBookings = allBookings.filter(attendee_id=attendee.id)
                buddies = ''
                hotelname = ''
                hotelLocation = ''
                hotelBeds = ''
                hotelCost = ''

                if actualBuddyName == '':
                    for booking in attendeeBookings:
                        buddyList = reqBuddy.filter(booking_id=booking.id)
                        for buddy in buddyList:
                            if buddy.name:
                                buddies = buddies + buddy.name
                            elif buddy.buddy_id:
                                buddies = buddies + buddy.buddy.firstname + " " + buddy.buddy.lastname
                            else:
                                buddies = buddies + ""

                        matchId = matchLines.filter(booking_id=booking.id)

                        if matchId:
                            actualBuddyList = matchLines.filter(match_id=matchId[0].match_id)
                            for actualBuddy in actualBuddyList:
                                if actualBuddy.booking_id != booking.id:
                                    actualBuddyName = actualBuddyName + actualBuddy.booking.attendee.firstname + " " + actualBuddy.booking.attendee.lastname
                                    hotelname = actualBuddy.booking.room.hotel.name
                                    hotelLocation = actualBuddy.booking.room.hotel.location.name
                                    hotelBeds = actualBuddy.booking.room.beds
                                    hotelCost = actualBuddy.booking.room.cost

                attRow.append(hotelname)
                attRow.append(hotelLocation)
                attRow.append(buddies)
                attRow.append(actualBuddyName)
                attRow.append(hotelBeds)
                attRow.append(hotelCost)

            ws.append(attRow)

        if not os.path.exists("attendeeList"):
            os.makedirs("attendeeList")

        wb.save("attendeeList/AllAttendee.xlsx")
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="AllAttendee.xls"'
        return response

    def set_export_order(request):
        response_data = {}
        if EventView.check_permissions(request, 'export_filter_permission'):
            exports_order = json.loads(request.POST.get('filter_export_order'))
            for export in exports_order:
                ExportRule.objects.filter(id=export['filter_export_id']).update(export_order=export['order'])
            response_data['success'] = 'Export Filters Ordered Successfully'
        else:
            response_data['error'] = 'You do not have Permission to do this'
        return HttpResponse(json.dumps(response_data), content_type="application/json")

    def get_export_order(group_id):
        export = ExportRule.objects.values('export_order').filter(group_id=group_id).aggregate(Max('export_order'))
        if export['export_order__max']:
            export_order = export['export_order__max'] + 1
        else:
            export_order = 1
        return export_order

    def question_history(question_history, new_question_answer, userId):
        question_answer_info = {}
        x_answer = Answers.objects.filter(question_id=new_question_answer['question_id'], user_id=userId)
        if x_answer:
            question_answer_info['old_data'] = str(x_answer[0].value).strip()
            if type(new_question_answer['answer']) is datetime.time:
                question_answer_info['new_data'] = str(new_question_answer['answer'])
            else:
                question_answer_info['new_data'] = str(new_question_answer['answer']).strip()
            if question_answer_info['old_data'] != question_answer_info['new_data']:
                question_answer_info['id'] = new_question_answer['question_id']
                question_answer_info['name'] = x_answer[0].question.title.strip()
                question_history.append(question_answer_info)

        else:
            if type(new_question_answer['answer']) is datetime.time:
                question_answer_info['new_data'] = str(new_question_answer['answer'])
            else:
                question_answer_info['new_data'] = str(new_question_answer['answer']).strip()
            question_answer_info['id'] = new_question_answer['question_id']
            question_answer_info['old_data'] = "Old Data Not Found"
            question = Questions.objects.get(pk=new_question_answer['question_id'])
            question_answer_info['name'] = question.title.strip()
            question_history.append(question_answer_info)
        return question_history

    def travel_history(travel_history, new_travel, userId):
        travel_info = {}
        x_travel = TravelAttendee.objects.filter(travel_id=new_travel['travel_id'], attendee_id=userId)
        if x_travel:
            travel_info['old_data'] = x_travel[0].status.strip()
            travel_info['new_data'] = new_travel['status'].strip()
            if travel_info['old_data'] != travel_info['new_data']:
                travel_info['id'] = new_travel['travel_id']
                travel_info['name'] = x_travel[0].travel.name.strip()
                travel_history.append(travel_info)
        return travel_history

    def session_history(session_history, new_session, userId):
        session_info = {}
        x_session = SeminarsUsers.objects.filter(session_id=new_session['session_id'], attendee_id=userId)
        if x_session:
            session_info['old_data'] = x_session[0].status.strip()
            session_info['new_data'] = new_session['status'].strip()
            if session_info['old_data'] != session_info['new_data']:
                session_info['id'] = new_session['session_id']
                session_info['name'] = x_session[0].session.name.strip()
                session_history.append(session_info)
        else:
            session_info['id'] = new_session['session_id']
            ssn = Session.objects.get(pk=new_session['session_id'])
            session_info['name'] = ssn.name.strip()
            session_info['old_data'] = "Old Data Not found"
            session_info['new_data'] = new_session['status'].strip()
            session_history.append(session_info)
        return session_history

    def approve_import_list(request):
        from_ten_days = datetime.today() - timedelta(days=10)
        ImportChangeRequest.objects.filter(created_at__lte=from_ten_days).exclude(status=1).delete()

        event_id = request.session['event_auth_user']['event_id']
        newlist = ImportChangeRequest.objects.filter(status=1, event_id=event_id)

        for item in newlist:
            item.created_at = TimeDetailView.utc_to_local(request, str(item.created_at))
        context = {
            'items': newlist
        }
        return render(request, 'export-import/approve_list.html', context)

    def delete_approve_import_item(request):
        id = request.POST.get('id')
        approve_item = ImportChangeRequest.objects.filter(id=id)
        if approve_item:
            approve_item.update(status=4)
            result = {'success': 'Successfully Deleted'}
        else:
            result = {'error': 'Not Found'}
        return HttpResponse(json.dumps(result), content_type='application/json')

    def view_approve_item(request, pk):
        newlist = ImportChangeRequest.objects.filter(id=pk)
        context = {
            'items': {'cd': json.loads(newlist[0].changed_data), 'id': newlist[0].id}
        }
        return render(request, 'export-import/approve_import.html', context)

    def check_import_status(request):
        event_id = request.session['event_auth_user']['event_id']
        imports = ImportChangeRequest.objects.filter(status=1,event_id=event_id)
        if imports:
            return HttpResponse(json.dumps({'error': "Approve Import first"}))
        else:
            return HttpResponse(json.dumps({'success': "Active import"}))

    def save_import(request):
        attendee_data_elements = []
        current_import = ImportChangeRequest.objects.get(pk=request.POST.get('id'))
        if current_import and current_import.status == 1:
            changing_elements = json.loads(request.POST.get('data'))
            while len(changing_elements) > 0:
                single_attendee_data = []
                single_attendee_id = changing_elements[0]['attendee_id']
                to_remove = 0
                for item in changing_elements:
                    if single_attendee_id == item['attendee_id']:
                        single_attendee_data.append(item)
                        to_remove += 1
                    else:
                        break
                del changing_elements[0:to_remove]
                attendee_data_elements.append(single_attendee_data)

            task = threading.Thread(target=ExcelView.save_function, args=(request, attendee_data_elements))
            task.start()

            current_import.status = 0
            current_import.save()
        response = dict(result=True, message="Saving import in progress...")
        return JsonResponse(response)

    saved_attendee_counter = 0
    total_import_attendee = 0

    def save_function(request, attendee_data_elements):
        new_attendees_id = []
        update_attendees_id = []
        ExcelView.saved_attendee_counter = 0
        ExcelView.total_import_attendee = len(attendee_data_elements)
        admin_id = request.session['event_auth_user']['id']
        event_id = request.session['event_auth_user']['event_id']
        data_name = request.POST.get('data_name')
        turn_order_pending = "make_pending" in data_name
        try:
            new_attendees = []
            for attendee_data_element in attendee_data_elements:
                with transaction.atomic():
                    att_id = int(attendee_data_element[0]['attendee_id'])
                    if att_id < 0:
                        new_att = dict(att=att_id, data=dict(g=[], q=[], s=[], t=[], h=[]))
                        for element in attendee_data_element:
                            new_att = ExcelView.new_att_info_insert(new_att, element)
                        new_attendees.append(new_att)
                    elif att_id > 0:
                        for element in attendee_data_element:
                            if element['attribute_name'] == 'g':
                                ExcelView.save_general_from_import(request, element['attribute_id'], int(element['attendee_id']), str(element['new_value'].strip()))
                            elif element['attribute_name'] == 'q':
                                ExcelView.save_answer_by_question_attendee_id(request, element['attribute_id'], element['attendee_id'], element['new_value'])
                            elif element['attribute_name'] == 's':
                                ExcelView.save_session_status(request, element['attribute_id'], element['attendee_id'], element['new_value'])
                            elif element['attribute_name'] == 't':
                                ExcelView.save_travel_status(request, element['attribute_id'], element['attendee_id'], element['new_value'])
                            elif element['attribute_name'] == 'h':
                                if int(element['attendee_id']) > 0:
                                    ExcelView.save_hotel_update(request, element['attendee_id'], element['new_value'], admin_id, event_id)
                        update_attendees_id.append(att_id)
                        if turn_order_pending:
                            open_order = EconomyLibrary.get_open_order_by_attendee(att_id)
                            if open_order:
                                EconomyLibrary.change_order_status(open_order.get('order_number'), 'pending', event_id, att_id, admin_id)
                        ExcelView.saved_attendee_counter += 1
            if new_attendees:
                new_attendees_id = ExcelView.save_new_attendees(request, new_attendees, turn_order_pending)

            if "send_mail" in data_name:
                ExcelView.import_send_mail(request, new_attendees_id, update_attendees_id)
        except Exception as excp:
            ErrorR.efail(excp)

    def get_import_state(request):
        response = dict(complete=False, percentage=0)
        if ExcelView.total_import_attendee == ExcelView.saved_attendee_counter:
            response['complete'] = True
        else:
            response['percentage'] = round((ExcelView.saved_attendee_counter * 100) / ExcelView.total_import_attendee)
        return JsonResponse(response)

    def import_send_mail(request, add_att_ids, edit_att_ids):
        email_activities = []
        for att_id in add_att_ids:
            attendee = Attendee.objects.get(id=att_id)
            add_email_settings = Setting.objects.filter(name='attendee_add_confirmation',event_id=attendee.event_id)
            if add_email_settings.exists():
                email_content = EmailContents.objects.filter(id=int(add_email_settings[0].value))  # there will be different email content depend on add attendee or registration
                if email_content.exists():
                    email_content = email_content[0]
                    EmailView.add_or_update_email_receivers(attendee, email_content.id,
                                                            request.session['event_auth_user']['id'])
                    message_history = MessageHistory(subject=email_content.subject, message='N/A',
                                                     admin_id=request.session['event_auth_user']['id'],
                                                     type='mail')
                    message_history.save()
                    activity_history = ActivityHistory(attendee_id=attendee.id,
                                                       admin_id=request.session['event_auth_user'][
                                                           'id'],
                                                       activity_type='message', category='message',
                                                       message_id=message_history.id,
                                                       event_id=request.session['event_auth_user'][
                                                           'event_id'])
                    email_activities.append(activity_history)
                    EmailView.send_email(request,attendee,email_content)
        for att_id in edit_att_ids:
            attendee = Attendee.objects.get(id=att_id)
            edit_email_settings = Setting.objects.filter(name='attendee_edit_confirmation',event_id=attendee.event_id)
            if edit_email_settings.exists():
                email_content = EmailContents.objects.filter(id=int(edit_email_settings[0].value))  # there will be different email content depend on update attendee or registration
                if email_content.exists():
                    email_content = email_content[0]
                    EmailView.add_or_update_email_receivers(attendee, email_content.id,
                                                            request.session['event_auth_user']['id'])
                    message_history = MessageHistory(subject=email_content.subject, message='N/A',
                                                     admin_id=request.session['event_auth_user']['id'],
                                                     type='mail')
                    message_history.save()
                    activity_history = ActivityHistory(attendee_id=attendee.id,
                                                       admin_id=request.session['event_auth_user'][
                                                           'id'],
                                                       activity_type='message', category='message',
                                                       message_id=message_history.id,
                                                       event_id=request.session['event_auth_user'][
                                                           'event_id'])
                    email_activities.append(activity_history)
                    EmailView.send_email(request,attendee,email_content)
        ActivityHistory.objects.bulk_create(email_activities)

    def save_new_attendees(request, new_attendees, turn_order_pending):
        att_id_for_list = []
        admin_id = request.session['event_auth_user']['id']
        event_id = request.session['event_auth_user']['event_id']
        setting_uid_length = Setting.objects.filter(name='uid_length', event_id=event_id)
        if setting_uid_length.exists():
            uid_length = int(setting_uid_length[0].value)
        else:
            uid_length = 16
        current_language = LanguageView.get_current_preset(request)
        # all_activities = []
        for new_att in new_attendees:
            try:
                with transaction.atomic():
                    # tt = time.time()
                    email = ""
                    fname = ""
                    lname = ""
                    phone = ""
                    for q in new_att["data"]["q"]:
                        if q["defination"] == "email":
                            email = q["value"]
                        elif q["defination"] == "firstname":
                            fname = q["value"]
                        elif q["defination"] == "lastname":
                            lname = q["value"]
                        elif q["defination"] == "phone":
                            phone = q["value"]

                    if len(email) < 1 or len(fname) < 1 or len(lname) < 1:
                        ExcelView.saved_attendee_counter += 1
                        continue

                    chars = string.ascii_uppercase + string.digits
                    size = 6
                    password = ''.join(random.choice(chars) for _ in range(size))

                    flag = True
                    while flag:
                        secret_key = ''.join(
                            random.choice(string.ascii_lowercase + string.digits + string.ascii_uppercase) for _ in range(uid_length))
                        checkUniquity = Attendee.objects.filter(secret_key__contains=secret_key).count()
                        if checkUniquity < 1:
                            flag = False

                    flag = True
                    while flag:
                        badge_key = ''.join(
                            random.choice(string.ascii_lowercase + string.digits + string.ascii_uppercase) for _ in range(uid_length))
                        checkUniquity = Attendee.objects.filter(bid__contains=badge_key).count()
                        if checkUniquity < 1:
                            flag = False

                    att = Attendee(firstname=fname, lastname=lname, email=email, phonenumber=phone,
                                   event_id=event_id, secret_key=secret_key, bid=badge_key, password=make_password(password), language_id=current_language.preset_id)
                    att.save()
                    new_activity_history = ActivityHistory(attendee_id=att.id, admin_id=admin_id, activity_type='register', category='event', event_id=event_id)
                    new_activity_history.save()

                    for g in new_att["data"]["g"]:
                        ExcelView.save_general_from_import(request, g['id'], att.id, g['value'])
                    for q in new_att["data"]["q"]:
                        ExcelView.save_answer_by_question_attendee_id(request, int(q['id']), att.id, q['value'])
                    for s in new_att["data"]["s"]:
                        ExcelView.save_session_status(request, int(s['id']), att.id, s['value'])
                    for t in new_att["data"]["t"]:
                        ExcelView.save_travel_status(request, int(t['id']), att.id, t['value'])
                    for h in new_att["data"]["h"]:
                        ExcelView.save_hotel_update(request, att.id, h['value'], admin_id, event_id)

                    att_id_for_list.append(att.id)
                    if turn_order_pending:
                        open_order = EconomyLibrary.get_open_order_by_attendee(att.id)
                        if open_order:
                            EconomyLibrary.change_order_status(open_order.get('order_number'), 'pending', event_id, att.id, admin_id)
            except :
                pass
            ExcelView.saved_attendee_counter += 1

        return att_id_for_list

    def new_att_info_insert(temp_att, element):
        if element['attribute_name'] == 'g':
            temp_att["data"]["g"].append(
                {"id": element['attribute_id'], "value": str(element['new_value'].strip())})
        elif element['attribute_name'] == 'q':
            temp_att["data"]["q"].append(
                {"id": element['attribute_id'], "value": str(element['new_value'].strip()),
                 "defination": element['defination']})
        elif element['attribute_name'] == 's':
            temp_att["data"]["s"].append(
                {"id": element['attribute_id'], "value": str(element['new_value'].strip())})
        elif element['attribute_name'] == 't':
            temp_att["data"]["t"].append(
                {"id": element['attribute_id'], "value": str(element['new_value'].strip())})
        elif element['attribute_name'] == 'h':
            temp_att["data"]["h"].append(
                {"id": element['attribute_id'], "value": str(element['new_value'].strip())})
        return temp_att

    def testQuestion(value, question_id):
        question = Questions.objects.get(id=question_id)
        response_data = {}
        response_data['question'] = question.title
        if question:
            if question.type == 'text':
                if question.actual_definition == 'firstname' or question.actual_definition == 'lastname' or question.actual_definition == 'email' or question.actual_definition == 'phone':
                    if value == '':
                        response_data['valid'] = False
                        response_data['reason'] = str(question.title) + " can't be empty"
                    else:
                        response_data['valid'] = True
                    if question.actual_definition == 'email':
                        if ExcelView.validateEmail(value) == 0:
                            response_data['valid'] = False
                            response_data['reason'] = "Email is not in the Correct Format"
                else:
                    response_data['valid'] = True
            elif question.type == 'radio_button' or question.type == 'select' or question.type == 'checkbox':
                if value != '':
                    options = Option.objects.filter(question_id=question_id)
                    found = False
                    option_str = ''
                    for option in options:
                        if option_str != '':
                            option_str = option_str + ','
                        option_str = option_str + str(option.option)
                        if option.option == value:
                            found = True
                    if found:
                        response_data['valid'] = True
                    else:
                        response_data['valid'] = False
                        response_data['reason'] = str(
                            question.title) + "'s value is not valid. options are " + option_str
                        response_data['options'] = "options are " + option_str
                else:
                    response_data['valid'] = True

            else:
                response_data['valid'] = True

        return response_data

    def save_general_from_import(request, qid, aid, new_answer, all_activities=None):
        event_id = request.session['event_auth_user']['event_id']
        admin_id = request.session['event_auth_user']['id']
        if qid == "rdate":
            if aid > 0:
                new_answer = datetime.strptime(new_answer.split('.')[0], "%Y-%m-%d %H:%M:%S")
                Attendee.objects.filter(id=aid).update(created=new_answer)
        elif qid == "udate":
            if aid > 0:
                new_answer = datetime.strptime(new_answer.split('.')[0], "%Y-%m-%d %H:%M:%S")
                Attendee.objects.filter(id=aid).update(updated=new_answer)
        elif qid == "tag":
            if aid > 0:
                old_answer = ""
                if len(new_answer) > 0 and new_answer != "None":
                    old_tags = AttendeeTag.objects.filter(attendee_id=aid)
                    for o_t in old_tags:
                        old_answer = old_answer + o_t.tag.name + ', '
                    if len(old_answer) > 1:
                        old_answer = old_answer[:-2]
                    else:
                        old_answer = "Empty"

                    tags = new_answer.split(',')
                    temp_tags = []
                    for tag in tags:
                        t = Tag.objects.filter(name=tag, event_id=event_id)
                        if t.exists():
                            t = t[0]
                            temp_tags.append(t.id)
                            att = AttendeeTag.objects.filter(tag_id=t.id, attendee_id=aid)
                            if not att:
                                new_att_tag = AttendeeTag(tag_id=t.id, attendee_id=aid)
                                new_att_tag.save()
                        else:
                            new_tag_create = Tag(name=tag, event_id=event_id)
                            new_tag_create.save()
                            new_att_tag = AttendeeTag(tag_id=new_tag_create.id, attendee_id=aid)
                            new_att_tag.save()
                            temp_tags.append(new_tag_create.id)

                    AttendeeTag.objects.filter(attendee_id=aid).exclude(tag_id__in=temp_tags).delete()

                else:
                    new_answer = "Empty"
                    att_tags = AttendeeTag.objects.filter(attendee_id=aid)
                    for att_tag in att_tags:
                        old_answer = old_answer + att_tag.tag.name + ', '
                        att_tag.delete()
                    if len(old_answer) > 1:
                        old_answer = old_answer[:-2]

                new_activity = ActivityHistory(activity_type="update", category="tag", admin_id=admin_id,
                                               attendee_id=aid, old_value=old_answer,
                                               new_value=new_answer, event_id=event_id)
                if all_activities:
                    all_activities.append(new_activity)
                else:
                    new_activity.save()

        elif qid == "group":
            if aid > 0:
                old_answer = ""
                if len(new_answer) > 0 and new_answer != "None":
                    old_groups = AttendeeGroups.objects.filter(attendee_id=aid)
                    for o_g in old_groups:
                        old_answer = old_answer + o_g.group.name + ', '
                    if len(old_answer) > 1:
                        old_answer = old_answer[:-2]
                    else:
                        old_answer = "Empty"

                    groups = new_answer.split(',')
                    temp_groups = []
                    for group in groups:
                        g = Group.objects.filter(name=group, type="attendee", event_id=event_id).first()
                        if g:
                            temp_groups.append(g.id)
                            att = AttendeeGroups.objects.filter(group_id=g.id, attendee_id=aid)
                            if not att.exists():
                                new_att_group = AttendeeGroups(group_id=g.id, attendee_id=aid)
                                new_att_group.save()
                        else:
                            group_order = 1
                            group_max_order = Group.objects.filter(type="attendee", event_id=event_id).aggregate(Max('group_order'))
                            if group_max_order['group_order__max']:
                                group_order = group_max_order['group_order__max']

                            current_language_id = LanguageH.get_current_language_id(event_id)
                            group_lang = group.replace('"', "&quot;").replace("'", "&apos;")
                            attendee_group_lang = str({current_language_id: group_lang}).replace("'", '"')
                            new_group = Group(name=group, name_lang=attendee_group_lang, type='attendee', event_id=event_id, group_order=group_order)
                            new_group.save()
                            new_att_group = AttendeeGroups(group_id=new_group.id, attendee_id=aid)
                            new_att_group.save()
                            temp_groups.append(new_group.id)

                    AttendeeGroups.objects.filter(attendee_id=aid).exclude(group_id__in=temp_groups).delete()
                else:
                    new_answer = "Empty"
                    att_groups = AttendeeGroups.objects.filter(attendee_id=aid)
                    for att_group in att_groups:
                        old_answer = old_answer + att_group.group.name + ', '
                        att_group.delete()
                    if len(old_answer) > 1:
                        old_answer = old_answer[:-2]

                new_activity = ActivityHistory(activity_type="update", category="group", admin_id=admin_id,
                                               attendee_id=aid, old_value=old_answer,
                                               new_value=new_answer, event_id=event_id)
                if all_activities:
                    all_activities.append(new_activity)
                else:
                    new_activity.save()
        if all_activities:
            return all_activities

    def save_answer_by_question_attendee_id(request, qid, aid, new_answer, all_activities=None):
        count_row = 0
        question = Questions.objects.get(pk=qid)
        event_id = request.session['event_auth_user']['event_id']
        if question:
            ans = Answers.objects.filter(question_id=qid, user_id=aid)
            admin_id = request.session['event_auth_user']['id']
            if ans:
                if question.actual_definition == 'firstname':
                    attendee = Attendee.objects.filter(id=aid).update(firstname=new_answer)
                elif question.actual_definition == 'lastname':
                    attendee = Attendee.objects.filter(id=aid).update(lastname=new_answer)
                elif question.actual_definition == 'phone':
                    attendee = Attendee.objects.filter(id=aid).update(phonenumber=new_answer)
                elif question.actual_definition == 'email':
                    attendee = Attendee.objects.filter(id=aid).update(email=new_answer)
                old_ans = ans[0]
                if new_answer == '[REMOVE]' and ans[0].question.group.name != "Default":
                    ans.delete()
                    activity = ActivityHistory(activity_type="update", category="question", admin_id=admin_id,
                                               attendee_id=aid, question_id=qid, old_value=old_ans.value,
                                               new_value="Empty", event_id=event_id)
                    activity.save()

                else:
                    if ans[0].question.type == 'checkbox':
                        new_answer = new_answer.replace(',','<br>')
                    count_row = ans.update(value=new_answer)
                    activity = ActivityHistory(activity_type="update", category="question", admin_id=admin_id,
                                               attendee_id=aid, question_id=qid, old_value=old_ans.value,
                                               new_value=new_answer, event_id=event_id)
                    activity.save()
            else:
                if question.type == 'checkbox':
                    new_answer = new_answer.replace(',', '<br>')
                ans = Answers(user_id=aid, question_id=qid, value=new_answer)
                count_row = ans.save()

                activity = ActivityHistory(activity_type="update", category="question", admin_id=admin_id,
                                           attendee_id=aid, question_id=qid, old_value="Empty", new_value=ans.value,
                                           event_id=event_id)
                if all_activities:
                    all_activities.append(activity)
                else:
                    activity.save()

        if all_activities:
            return all_activities
        return count_row

    def save_travel_status(request, travel_id, aid, new_status, all_activities=None):
        count_row = 0
        if int(aid) > 0:
            travel = TravelAttendee.objects.filter(travel_id=travel_id, attendee_id=aid)
            if travel:
                if new_status == '[delete]':
                    travel.delete()
                else:
                    count_row = travel.update(value=new_status)
            else:
                travel = TravelAttendee(attendee_id=aid, travel_id=travel_id, status=new_status)
                count_row = travel.save()
        if all_activities:
            return all_activities
        return count_row

    def save_session_status(request, session_id, aid, new_status, all_activities=None):
        session = SeminarsUsers.objects.filter(session_id=session_id, attendee_id=aid).first()
        admin_id = request.session['event_auth_user']['id']
        event_id = request.session['event_auth_user']['event_id']
        if session:
            if new_status == '[REMOVE]':
                old_status = session.status
                session.delete()
                SessionDetail.notify_queue_user(event_id, session_id)
                activity = ActivityHistory(activity_type="update", category="session", attendee_id=aid,
                                           admin_id=admin_id, session_id=session_id, old_value=session.status,
                                           new_value="Deleted", event_id=event_id)
                activity.save()
                if old_status == 'attending':
                    ExcelView.economy_remove_for_import(request, 'session', session_id, aid)
            else:
                old_status = session.status
                session.status = new_status
                session.save()
                if new_status == "not-attending" and session.status == "attending":
                    SessionDetail.notify_queue_user(event_id, session_id)

                activity = ActivityHistory(activity_type="update", category="session", attendee_id=aid, admin_id=admin_id,
                                           session_id=session_id, old_value=old_status, new_value=new_status, event_id=event_id)
                activity.save()
                if new_status == 'attending':
                    ExcelView.economy_place_for_import(request, 'session', session_id, aid)
                elif old_status == 'attending':
                    ExcelView.economy_remove_for_import(request, 'session', session_id, aid)
        else:
            session = SeminarsUsers(attendee_id=aid, session_id=session_id, status=new_status)
            session.save()

            activity = ActivityHistory(activity_type="update", category="session", attendee_id=aid, admin_id=admin_id,
                                       session_id=session_id, old_value="Empty", new_value=new_status, event_id=event_id)
            if all_activities:
                all_activities.append(activity)
            else:
                activity.save()
            if new_status == 'attending':
                ExcelView.economy_place_for_import(request, 'session', session_id, aid)
        if all_activities:
            return all_activities
        return

    def save_hotel_update(request, aid, hotel_data, admin_id, event_id, all_activities=None):
        try:
            hotel_data = eval(hotel_data)
            cursor = connection.cursor()
            new_booking_id = None
            for item in hotel_data:
                booking_id = item.get('booking_id')
                if item['table'] == 'bookings':
                    if item['action'] == 'insert':
                        booking = Booking(attendee_id=aid, check_in=item['data']['check_in'], check_out=item['data']['check_out'], room_id=item['data']['room_id'])
                        booking.save()
                        new_booking_id = booking.id
                        ActivityHistory(attendee_id=aid, admin_id=admin_id, activity_type="register", category="room",
                                        room_id=item['data']['room_id'], event_id=event_id).save()
                        booking_check_in_date = datetime.strptime(item['data']['check_in'], '%Y-%m-%d')
                        booking_check_out_date = datetime.strptime(item['data']['check_out'], '%Y-%m-%d')
                        booking_day_count = (booking_check_out_date - booking_check_in_date).days
                        ExcelView.economy_place_for_import(request, 'hotel', item['data']['room_id'], aid, new_booking_id, booking_day_count)
                    elif item['action'] == 'update':
                        existing_booking = Booking.objects.get(id=booking_id)
                        room_id = existing_booking.room_id
                        Booking.objects.filter(id=booking_id).update(check_in=item['data']['check_in'], check_out=item['data']['check_out'])
                        ExcelView.economy_update_hotel_cost(request, aid, booking_id, item['data']['check_in'], item['data']['check_out'], room_id, existing_booking.check_in, existing_booking.check_out)
                    elif item['action'] == 'delete':
                        del_booking = Booking.objects.get(id=booking_id)
                        booking_allotment_dates = [del_booking.check_in + timedelta(n) for n in range(0, (del_booking.check_out - del_booking.check_in).days)]
                        del_booking.delete()
                        ActivityHistory(attendee_id=aid, admin_id=admin_id, activity_type="delete", category="room",
                                        room_id=item['data']['room_id'], event_id=event_id).save()
                        ExcelView.economy_remove_for_import(request, 'hotel', item['data']['room_id'], aid, booking_id, booking_allotment_dates)
                elif item['table'] == 'requested_buddies':
                    if item['action'] == 'insert':
                        inserted_booking_id = new_booking_id if new_booking_id else booking_id
                        if item['data']['user_exists']:
                            RequestedBuddy(booking_id=inserted_booking_id, buddy_id=item['data']['attendee_id']).save()
                        else:
                            RequestedBuddy(booking_id=inserted_booking_id, exists=False, email=item['data']['email']).save()
                    elif item['action'] == 'delete':
                        RequestedBuddy.objects.filter(booking_id=booking_id).delete()
                elif item['table'] == 'match_line':
                    if item['action'] == 'delete':
                        MatchLine.objects.filter(match_id__in=item['data']['match_ids']).delete()
                elif item['table'] == 'matches':
                    if item['action'] == 'delete':
                        Match.objects.filter(id__in=item['data']['match_ids']).delete()
                elif item['table'] == 'actual_buddy_block':
                    # here we process actual buddy things
                    HotelLamda.handle_actual_buddy(request, new_booking_id if new_booking_id else booking_id, item['data'], event_id, cursor)
        except Exception as ex:
            ErrorR.c_red('XXX')
            ErrorR.efail(ex)

    def economy_place_for_import(request, item_type, item_id, attendee_id, booking_id=None, booking_day_count=None):
        admin_id = request.session['event_auth_user']['id']
        event_id = request.session['event_auth_user']['event_id']
        EconomyLibrary.place_order(event_id, attendee_id, item_type, item_id, admin_id, booking_id=booking_id, booking_day_count=booking_day_count)
        return

    def economy_remove_for_import(request, item_type, item_id, attendee_id, booking_id=None, booking_allotment_dates=None):
        admin_id = request.session['event_auth_user']['id']
        event_id = request.session['event_auth_user']['event_id']
        order_info = EconomyLibrary.get_order_id(attendee_id, item_type, item_id, booking_id)
        if order_info:
            EconomyLibrary.remove_item_from_order(event_id, attendee_id, order_info['order_id'], item_id, booking_id, admin_id, booking_allotment_dates)
        return

    def economy_update_hotel_cost(request, attendee_id, booking_id, cin, cout, room_id, old_check_in, old_check_out):
        admin_id = request.session['event_auth_user']['id']
        event_id = request.session['event_auth_user']['event_id']
        previous_day_count = (old_check_out - old_check_in).days
        check_in = datetime.strptime(cin, "%Y-%m-%d")
        check_out = datetime.strptime(cout, "%Y-%m-%d")
        new_day_count = (check_out - check_in).days
        old_booking_dates = ExcelView.get_date_list(old_check_in, old_check_out)
        new_booking_dates = ExcelView.get_date_list(check_in, check_out)
        new_extra_booking_dates = list(set(new_booking_dates) - set(old_booking_dates))
        if previous_day_count != new_day_count:
            day_difference = 0
            if new_day_count > previous_day_count:
                day_difference = new_day_count - previous_day_count

            EconomyLibrary.update_hotel_cost(event_id, attendee_id, room_id, new_day_count, day_difference, booking_id, new_booking_id=None, new_extra_booking_dates=new_extra_booking_dates, admin_id=admin_id)
        elif old_check_in != check_in or old_check_out != check_out:
            EconomyLibrary.update_hotel_for_allotment(event_id, attendee_id, room_id, booking_id, old_booking_dates, new_booking_dates, admin_id)

    def get_date_list(check_in, check_out):
        return [check_in + timedelta(n) for n in range(0, (check_out - check_in).days)]

    def import_action(request):
        if request.POST.get('action') == "Cancel":
            ImportChangeRequest.objects.filter(id=request.POST.get('icr_id')).delete()
        elif request.POST.get('action') == "Proceed":
            ignored_attendees = request.POST.get("ignored_attendees")
            ignored_attendees = json.loads(ignored_attendees) if ignored_attendees else []
            if ignored_attendees:
                ignored_attendees = [int(item) for item in ignored_attendees]
                icr_id = request.POST.get("icr_id")
                icr = ImportChangeRequest.objects.get(id=icr_id)
                changed_data_original = json.loads(icr.changed_data)
                changed_data = json.loads(icr.changed_data)
                for cd in changed_data_original:
                    if cd["Attendee"] in ignored_attendees:
                        changed_data.remove(cd)
                icr.changed_data = json.dumps(changed_data)
                icr.save()
        else:
            ImportChangeRequest.objects.filter(id=request.POST.get('icr_id')).update(status=1)
            ImportChangeStatus.objects.filter(id=request.POST.get('import_change_status_id')).update(status=1)
            return redirect("render_import_change_page", pk=request.POST.get("import_change_status_id"))
        return redirect('import-list')

    def check_new_attendee_duplicate_email(request):

        return

    def hotel_import(request):
        Uploader.handle_uploaded_file(request.FILES.get('upload_file'), 'hotel_import.xlsx')

        from openpyxl import load_workbook

        wb = load_workbook(filename='attendeeList/hotel_import.xlsx', read_only=True)
        ws = wb['Attendees']  # ws is now an IterableWorksheet

        response_data = []
        # questionHeaders = []
        header_row = []

        # i=1
        item_row = 0
        for row in ws.rows:

            answers = []

            if item_row < 2:
                if item_row == 0:
                    for item in row:
                        item_val = item.value.split('-')
                        header_row.append(item_val)
                questionHeaders = row
                item_row = item_row + 1
                continue
            item_row = item_row + 1

            success_message = []
            question_history = []

            try:
                att = None
                if row[0].value == "att-uid":
                    attendee_id = row[0].value
                    att = Attendee.objects.filter(id=attendee_id)
                else:
                    email_index = Uploader.get_email_index(header_row)
                    if email_index > 0:
                        att = Uploader.get_attendee_by_email(row[email_index].value)
                        attendee_id = att.id
                if not att:
                    continue

                content_serial = 0
                for item in row:
                    if header_row[content_serial][0] == "att":
                        if header_row[content_serial][1] == 'rdate':
                            registration_date = item.value

                    elif header_row[content_serial][0] == "q":
                        answers.extend(
                            [{"question_id": header_row[content_serial][1], "answer": item.value, "a_type": 'text'}])

                    content_serial = content_serial + 1

                att_change_flag = False
                if att.exists():
                    att = att[0]

                    for answer in answers:
                        if answer['answer']:
                            if (Questions.objects.filter(id=answer['question_id']).exists()):
                                question_history = ExcelView.question_history(question_history, answer, attendee_id)

                    changed_data = {}
                    if len(question_history) > 0:
                        changed_data['questions'] = question_history
                    if len(changed_data) > 0:
                        response_data.append({"Attendee": attendee_id, 'data': changed_data})

            except Exception as e:
                response_data.append({"Attendee": "Error!!! =" + str(e)})

    def getTimezoneNow(request):
        setting_timezone = Setting.objects.filter(name='timezone',
                                                  event_id=request.session['event_auth_user']['event_id'])
        if setting_timezone:
            tzname = setting_timezone[0].value
            timezone_active = timezone(tzname)
            now = datetime.now(timezone_active)
            return now


class Uploader():
    def get_checksum(row):
        import hashlib
        hash_object = hashlib.md5(bytes(json.dumps(row), encoding='utf-8'))
        return hash_object.hexdigest()

    def import_complete_status(request):
        import_statusID = int(request.POST.get('import_status_id'))
        import_change_status = ImportChangeStatus.objects.filter(id=import_statusID).first()
        status = import_change_status.status
        response = {
            'status': status
        }
        return HttpResponse(json.dumps(response), content_type="application/json")

    def render_import_status(request, pk):
        context = {}
        import_change_status = ImportChangeStatus.objects.filter(id=pk).first()
        # context = dict(header_error=True, errors=[dict(incorrect_header='atdf-sdlf',available_headers='att-grp, att-tag',message='ok'), dict(incorrect_header='q-ssd', available_headers='q-question_id', message='ok')],
        # icr=ImportChangeRequest.objects.filter(id=import_change_status.import_change_id).first())
        if not import_change_status:
            return render(request, 'export-import/import-result.html', dict(not_found=True))
        if import_change_status.status == 1:
            if import_change_status.duplicate_attendees and json.loads(import_change_status.duplicate_attendees):
                deciding_attendees = json.loads(import_change_status.duplicate_attendees)
                deciding_attendees = sorted(deciding_attendees, key=lambda item: item["email"])
                deciding_attendee_list = []
                ignored_attendees = []
                email_checker = ""
                list_counter = -1
                for attendee in deciding_attendees:
                    if attendee["email"] != email_checker:
                        deciding_attendee_list.append([attendee])
                        list_counter += 1
                    else:
                        deciding_attendee_list[list_counter].append(attendee)
                    email_checker = attendee["email"]
                    ignored_attendees.append(attendee["id"])

                context["icr"] = ImportChangeRequest.objects.filter(id=import_change_status.import_change_id).first()
                context["deciding_attendees"] = deciding_attendee_list
                context["ignored_attendees"] = ignored_attendees
            else:
                return redirect('import-list')
        elif import_change_status.status == 2:
            context = {"errors": json.loads(import_change_status.message),
                       'icr': ImportChangeRequest.objects.filter(id=import_change_status.import_change_id).first(),
                       'import_change_status_id': import_change_status.id
                       }
        elif import_change_status.status == 3:
            context['success'] = "Not found any importing changes"

        elif import_change_status.status == 5:
            # for hotel import
            context_messages = json.loads(ImportChangeRequest.objects.filter(id=import_change_status.import_change_id)[0].changed_data)
            index = 0
            for ctx_msgs in context_messages:
                context_booking = ctx_msgs['booking']
                if isinstance(context_booking, int):
                    context_booking = Booking.objects.filter(id=context_booking)
                    if context_booking:
                        context_messages[index]['booking'] = context_booking[0]
                index += 1
            return render(request, 'export-import/import_hotel_result.html', {'messages': context_messages})
        elif import_change_status.status == 6:
            context['icr'] = ImportChangeRequest.objects.filter(id=import_change_status.import_change_id).first()
            context['header_error'] = True
            context['errors'] = json.loads(import_change_status.message)
        else:
            context['success'] = "Still in progress"

        return render(request, 'export-import/import-result.html', context)

    def upload(request):
        admin_id = request.session['event_auth_user']['id']
        event_id = request.session['event_auth_user']['event_id']

        unique_time = datetime.strftime(datetime.now(), "%Y-%m-%dT%H:%M:%S")
        filename = 'import_test/sampleimport_' + str(admin_id) + '_' + str(event_id) + '_' + unique_time + '.xlsx'
        import_change_status = ImportChangeStatus(filename=filename, status=0)
        import_change_status.save()

        if os.environ['ENVIRONMENT_TYPE'] == 'development':
            LambdaUpload.lambda_upload(request, filename)  # For Local
        else:
            Uploader.handle_uploaded_file(request.FILES.get('upload_file'), filename)  # For dev / stag / prod

        response = {
            'id': import_change_status.id
        }
        return HttpResponse(json.dumps(response), content_type="application/json")
        # return HttpResponse("Request in progress")

        from openpyxl import load_workbook

        wb = load_workbook(filename='attendeeList/sample_import.xlsx', read_only=True)
        ws = wb.worksheets[0]  # ws is now an IterableWorksheet

        response_data2 = []
        header_row = []

        # i=1
        item_row = 0
        error_flag = False
        message = []
        for row in ws.rows:

            answers = []
            travels = []
            sessions = []
            only_answers = []
            att_only_answers = []

            if item_row < 2:
                if row[0].value == "booking":
                    return HotelImport.dynamic_hotel_import(request, ws.rows)
                if item_row == 0:
                    for item in row:
                        if item.value is not None:
                            item_val = item.value.split('-')
                            header_row.append(item_val)
                item_row = item_row + 1
                continue
            item_row = item_row + 1

            question_history = []
            session_history = []
            travel_history = []

            try:
                att = None
                exception_identifier = ""
                if header_row[0][1] == "uid":
                    attendee_id = row[0].value
                    exception_identifier = str(attendee_id)
                    att = Attendee.objects.filter(id=attendee_id)
                else:
                    email_index = Uploader.get_email_index(header_row)
                    if email_index > 0:
                        att = Uploader.get_attendee_by_email(row[email_index].value)
                        attendee_id = att.id
                    exception_identifier = " row " + str(item_row)
                if not att:
                    continue

                row_items_checksum = []
                for item in row:
                    if item.value == None:
                        row_items_checksum.append("")
                    else:
                        row_items_checksum.append(str(item.value))

                row_items_checksum.pop(1)
                row_items_checksum.pop(2)
                checksum = Uploader.get_checksum(row_items_checksum)
                att = Attendee.objects.filter(id=attendee_id).first()

                if att.checksum == checksum:
                    att.checksum_flag = False
                    att.save()
                    continue

                if row[1].value == None:
                    att.checksum_flag = False

                elif row[1].value.lower().strip() == "no":
                    att.checksum = ""
                    att.checksum_flag = False

                elif row[1].value.lower().strip() == "yes":
                    att.checksum = checksum
                    att.checksum_flag = True

                att.save()

                content_serial = 0
                for item in row:

                    if len(header_row) <= content_serial:
                        content_serial += 1
                        continue
                    if header_row[content_serial][0] == "att":
                        if header_row[content_serial][1] == 'rdate':
                            registration_date = item.value

                    elif header_row[content_serial][0] == "q":
                        answers.extend(
                            [{"question_id": header_row[content_serial][1], "answer": item.value, "a_type": 'text'}])
                        if item.value:
                            only_answers.append(str(item.value))

                            att_only_answer = Answers.objects.filter(user_id=attendee_id,
                                                                     question_id=header_row[content_serial][1]).first()

                            if att_only_answer:
                                att_only_answers.append(att_only_answer.value)
                            else:
                                att_only_answers.append('')

                    elif header_row[content_serial][0] == "travel":
                        if str.lower(str(item.value)) == "attending" or str.lower(
                                str(item.value)) == "not-attending" or str.lower(
                            str(item.value)) == "in-queue" or str.lower(str(item.value)) == "deciding":
                            travels.extend(
                                [{"travel_id": header_row[content_serial][1], "status": str.lower(item.value)}])

                    elif header_row[content_serial][0].lower() == "session":
                        if str.lower(str(item.value)) == "attending" or str.lower(
                                str(item.value)) == "not-attending" or str.lower(
                            str(item.value)) == "in-queue" or str.lower(str(item.value)) == "deciding":
                            sessions.extend(
                                [{"session_id": header_row[content_serial][1], "status": str.lower(item.value)}])
                    content_serial = content_serial + 1

                att_change_flag = False
                if att:
                    for answer in answers:
                        if answer['answer']:
                            question = Questions.objects.filter(id=answer['question_id'])
                            if (question.exists()):
                                test_question = ExcelView.testQuestion(str(answer['answer']),
                                                                       int(answer['question_id']))

                                if test_question['valid'] == False:
                                    message.append([{'name': test_question['question'], 'type': 'Question',
                                                     'attendee': attendee_id, 'reason': test_question['reason']}])
                                    error_flag = True
                                else:
                                    question_history = ExcelView.question_history(question_history, answer, attendee_id)
                    for travel in travels:
                        if travel['travel_id']:
                            if (Travel.objects.filter(id=travel['travel_id']).exists()):
                                travel_history = ExcelView.travel_history(travel_history, travel, attendee_id)

                    for ssn in sessions:
                        if ssn['session_id']:
                            if (Session.objects.filter(id=ssn['session_id']).exists()):
                                test_session = General.testSession(attendee_id, int(ssn['session_id']))
                                if test_session['valid'] == False:
                                    message.append([{'name': test_session['session'], 'type': 'Session',
                                                     'attendee': attendee_id, 'reason': test_session['reason']}])
                                    error_flag = True
                                else:
                                    session_history = ExcelView.session_history(session_history, ssn, attendee_id)
                    changed_data = {}
                    if len(question_history) > 0:
                        changed_data['questions'] = question_history
                    if len(session_history) > 0:
                        changed_data['sessions'] = session_history
                    if len(travel_history) > 0:
                        changed_data['travels'] = travel_history
                    if len(changed_data) > 0:
                        response_data2.append({"Attendee": attendee_id, 'data': changed_data})
            except Exception as e:
                import sys
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print(exc_type, fname, exc_tb.tb_lineno)
                response_data2.append({"Attendee": "Error!!! =" + str(e)})
                message.append([{'question': '', 'attendee': exception_identifier, 'reason': "Error!!! =" + str(e)}])
                error_flag = True

        context = {
            'errors': message
        }
        admin_id = request.session['event_auth_user']['id']
        event_id = request.session['event_auth_user']['event_id']
        if len(response_data2) > 0:
            if error_flag == False:
                icr = ImportChangeRequest(event_id=event_id, changed_data=json.dumps(response_data2),
                                          imported_by_id=admin_id, status=1, created_at=datetime.now())
                icr.save()
                context = {
                    'success': "Success"
                }
            else:
                icr = ImportChangeRequest(event_id=event_id, changed_data=json.dumps(response_data2),
                                          imported_by_id=admin_id, status=3, created_at=datetime.now())
                icr.save()

            context['icr'] = icr

        else:
            context['success'] = "Not found any importing changes"

        return render(request, 'export-import/import-result.html', context)

    def handle_uploaded_file(f, filename):
        if os.environ['ENVIRONMENT_TYPE'] == 'development':
            ErrorR.c_red('locally import is not activated')
            raise Exception('locally import is not activated')

        # S3
        import boto
        from boto.s3.key import Key
        conn = boto.connect_s3(settings.AWS_ACCESS_KEY_ID, settings.AWS_SECRET_ACCESS_KEY, host=settings.AWS_STORAGE_HOST)
        bucket = conn.get_bucket(settings.AWS_STORAGE_BUCKET_NAME)
        filename_with_path = filename
        key_name = filename_with_path
        k = Key(bucket)
        k.key = key_name

        xl_file = io.BytesIO(f.read())

        if not k.exists():
            key = bucket.new_key(key_name)
            key.set_contents_from_string(xl_file.getvalue())
            key.set_metadata('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            key.set_acl('public-read')
            key.make_public()
        else:
            k.set_contents_from_string(xl_file.getvalue())
            k.set_metadata('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            k.set_acl('public-read')
            k.make_public()
            # END S3

    def get_email_index(rows):
        index = -1
        for row in rows:
            index += 1
            if row:
                if row[0] == "q":
                    question = Questions.objects.filter(id=row[1]).first()
                    if question.actual_definition == "email":
                        return index
        return -1

    def get_attendee_by_email(email):
        return Attendee.objects.filter(email=email).first()


class HotelImport:
    def dynamic_hotel_import(request, all_rows):
        messages = []
        item_row = -1
        erro_msg = []

        for rows in all_rows:
            with transaction.atomic():
                try:
                    update_msg = []
                    update_flag = False
                    item_row += 1
                    if item_row < 2:
                        continue

                    booking_id = rows[0].value
                    room_id = rows[6].value

                    x_booking = Booking.objects.filter(id=booking_id).first()
                    booking = x_booking
                    if booking:
                        x_room = booking.room
                        room = x_room
                        if x_room.id != room_id:
                            if room_id == "[REMOVE]":
                                x_booking.delete()
                                update_msg.append("Booking removed")
                                messages.append({'booking': booking, 'errors': erro_msg, 'msgs': update_msg})
                                continue
                            checkin_date = str(rows[9].value)
                            checkout_date = str(rows[10].value)
                            room_allotments = RoomAllotment.objects.filter(room_id=room_id, available_date__range=(
                                checkin_date, checkout_date))
                            is_available = False

                            if room_allotments:
                                for allotment in room_allotments:
                                    available = HotelImport.get_available_allotment(allotment, room_id)
                                    if available > 0:
                                        is_available = True
                                    else:
                                        is_available = False
                                        break
                            if is_available:
                                new_booking = Booking(attendee_id=x_booking.attendee_id, room_id=room_id,
                                                      check_in=rows[9].value, check_out=rows[10].value,
                                                      broken_up=x_booking.broken_up)
                                new_booking.save()
                                booking = new_booking
                                booking_id = booking.id
                                room = booking.room
                                x_booking.delete()
                                update_msg.append("Booking updated")
                                update_flag = True
                            else:
                                update_flag = True
                                erro_msg.append("Room not available")
                                continue
                        else:
                            if str(booking.check_in) != str(rows[9].value) or str(booking.check_out) != str(
                                    rows[10].value):

                                new_checkin_date = datetime.strptime(str(rows[9].value), "%Y-%m-%d").date()
                                new_checkout_date = datetime.strptime(str(rows[10].value), "%Y-%m-%d").date()

                                old_checkin_date = booking.check_in
                                old_checkout_date = booking.check_out

                                old_dates = []
                                new_dates = []

                                for single_date in HotelImport.daterange(new_checkin_date, new_checkout_date):
                                    new_dates.append(str(single_date))
                                for single_date in HotelImport.daterange(old_checkin_date, old_checkout_date):
                                    old_dates.append(str(single_date))

                                for dd in old_dates:
                                    if dd in new_dates:
                                        new_dates.remove(dd)

                                is_available = False
                                if len(new_dates) > 0:
                                    for new_date in new_dates:
                                        room_allotments = RoomAllotment.objects.filter(room_id=room_id,
                                                                                       available_date=str(new_date))
                                        for allotment in room_allotments:
                                            available = HotelImport.get_available_allotment(allotment, room_id)
                                            if available > 0:
                                                is_available = True
                                else:
                                    is_available = True

                                if is_available:
                                    from app.views.attendee_view import AttendeeView
                                    booking.check_in = str(rows[9].value)
                                    booking.check_out = str(rows[10].value)
                                    booking.save()
                                    get_match = MatchLine.objects.filter(booking_id=booking.id)
                                    booking_matches = []
                                    if get_match.exists():
                                        matches = MatchLine.objects.filter(match_id=get_match[0].match_id)
                                        for match in matches:
                                            booking_matches.append(match.booking_id)
                                    common_dates = AttendeeView.get_booking_common_dates(booking_matches)
                                    if len(common_dates) == 0:
                                        new_booking = Booking(attendee_id=booking.attendee_id, room_id=room_id,
                                                              check_in=rows[9].value, check_out=rows[10].value)
                                        new_booking.save()
                                        Booking.objects.get(id=booking.id).delete()
                                        update_msg.append("Booking created")
                                    elif len(common_dates) > 0:
                                        all_dates = []
                                        for my_date in common_dates:
                                            all_dates.append(str(my_date))
                                        end_date = max(common_dates) + timedelta(days=1)
                                        updated_match = Match.objects.filter(id=get_match[0].match_id).update(
                                            start_date=min(common_dates), end_date=end_date,
                                            all_dates=json.dumps(all_dates))
                                        update_msg.append("Booking updated")
                                    else:
                                        update_msg.append("Booking updated")
                                    update_flag = True
                                else:
                                    update_flag = True
                                    erro_msg.append("Room not available")

                        if room.description != rows[8].value:
                            room.description = rows[8].value
                            room.save()
                            update_flag = True
                            update_msg.append("Hotel Description Changed")

                        hotel = room.hotel
                        if hotel.name != rows[7].value:
                            hotel.name = rows[7].value
                            hotel.save()
                            update_flag = True
                            update_msg.append("Hotel name Changed")

                        if rows[11].value == "[REMOVE]":
                            RequestedBuddy.objects.filter(booking_id=booking.id).delete()
                            update_msg.append("Requested room buddy removed")
                            update_flag = True

                        if rows[12].value and rows[12].value.strip() != None:
                            update_msg.append(HotelImport.handle_actual_buddy(request, booking, rows[12].value))
                            update_flag = True

                        if update_flag:
                            messages.append({'booking': booking, 'errors': erro_msg, 'msgs': update_msg})

                    else:
                        erro_msg.append("Booking not found")
                        messages.append({'booking': {'id': booking_id}, 'errors': erro_msg, 'msgs': ''})


                except IOError as e:
                    messages.append(
                        {'booking': {'id': booking_id}, 'errors': ['Exception occurs: ' + str(e)], 'msgs': ''})
        return messages

    def get_available_allotment(allotment, room_id):
        from django.db.models.aggregates import Count
        matched_attendee = MatchLine.objects.values('match_id').annotate(count_match=Count('match_id')).filter(
            count_match__gt=1, match__room_id=room_id)
        count_matched_pairs = matched_attendee.filter(match__start_date__lte=allotment.available_date,
                                                      match__end_date__gt=allotment.available_date).count()

        match_id = []
        if matched_attendee.count() > 0:
            for match in matched_attendee:
                match_id.append(match['match_id'])
        count_matched_singles = MatchLine.objects.values('match_id').annotate(count_match=Count('match_id')).filter(
            match_id__in=match_id, booking__check_in__lte=allotment.available_date,
            booking__check_out__gt=allotment.available_date).exclude(match__start_date__lte=allotment.available_date,
                                                                     match__end_date__gt=allotment.available_date).count()
        matched_booking = []
        booking_matched = MatchLine.objects.filter(match_id__in=match_id)
        if booking_matched.exists():
            for booking in booking_matched:
                matched_booking.append(booking.booking_id)
        count_unmatched_attendee = Booking.objects.filter(room_id=room_id, check_in__lte=allotment.available_date,
                                                          check_out__gt=allotment.available_date).exclude(
            id__in=matched_booking).count()
        total = count_matched_pairs + count_matched_singles + count_unmatched_attendee
        available = allotment.allotments - total
        return available

    def daterange(start_date, end_date):
        for n in range(int((end_date - start_date).days)):
            yield start_date + timedelta(n)

    def handle_actual_buddy(request, attendee_booking, actual_buddy_list):
        results = []
        if actual_buddy_list == "[REMOVE]":
            my_match = MatchLine.objects.filter(booking_id=attendee_booking.id).delete()
            results.append(
                "Match delete for " + attendee_booking.attendee.firstname + " " + attendee_booking.attendee.lastname)
            return results
        result = {'success': False, 'message': 'Actual Buddy not updated'}

        actual_buddy_list = [x.strip() for x in actual_buddy_list.split(',')]
        all_pairable_bookings = [attendee_booking.id]
        for attendee_email in actual_buddy_list:
            attendee = Attendee.objects.filter(email=attendee_email).first()
            if attendee:
                first_date = attendee_booking.check_in
                last_date = attendee_booking.check_out
                my_booking = Booking.objects.filter(attendee_id=attendee.id)
                my_booking = my_booking.filter(Q(
                    Q(check_in__lte=first_date, check_out__gte=first_date) | Q(check_in__lte=last_date,
                                                                               check_out__gte=last_date) | Q(
                        check_in__gte=first_date, check_in__lte=last_date) | Q(check_out__gte=first_date,
                                                                               check_out__lte=last_date)) & Q(
                    room=attendee_booking.room))
                x_match = MatchLine.objects.filter(Q(booking_id__in=my_booking) | Q(booking_id=attendee_booking.id))
                x_matchlines_id = []
                for match in x_match:
                    x_matchlines_id.append(match.id)

                for actual_buddy_booking in my_booking:
                    all_pairable_bookings.append(actual_buddy_booking.id)
        if len(all_pairable_bookings) > 1:
            result = HotelView.pair_up_details(request, all_pairable_bookings)
        if result['success']:
            results.append(result['message'])
            MatchLine.objects.filter(id__in=x_matchlines_id).delete()
            results.append(str(len(x_matchlines_id)) + " matchlines deleted...")
        else:
            results.append(result['message'])
        return results
